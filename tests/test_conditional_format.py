"""Conditional formatting tests.

Rule XML is verified through openpyxl rather than by asserting on raw
strings, so these check the thing Excel actually reads.
"""
import os
import tempfile

import openpyxl
import pytest

from rvgsrust_xlsxwriter import (
    ConditionalFormat2ColorScale,
    ConditionalFormat3ColorScale,
    ConditionalFormatAverage,
    ConditionalFormatBlank,
    ConditionalFormatCell,
    ConditionalFormatDataBar,
    ConditionalFormatDate,
    ConditionalFormatDuplicate,
    ConditionalFormatError,
    ConditionalFormatFormula,
    ConditionalFormatText,
    ConditionalFormatTop,
    Format,
    Workbook,
)


def _apply(cf, first_row=0, first_col=0, last_row=4, last_col=0):
    """Write a small numeric column, attach cf to it, reload via openpyxl."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        wb = Workbook()
        ws = wb.add_worksheet()
        for i, value in enumerate([10, 20, 30, 40, 50]):
            ws.write(i, 0, value)
        ws.add_conditional_format(first_row, first_col, last_row, last_col, cf)
        wb.close(path)
        sheet = openpyxl.load_workbook(path).active
        return [r for cfmt in sheet.conditional_formatting for r in cfmt.rules]
    finally:
        if os.path.exists(path):
            os.remove(path)


def _ranges(cf):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        wb = Workbook()
        ws = wb.add_worksheet()
        ws.write(0, 0, 1)
        ws.add_conditional_format(0, 0, 4, 0, cf)
        wb.close(path)
        sheet = openpyxl.load_workbook(path).active
        return [str(c.sqref) for c in sheet.conditional_formatting]
    finally:
        if os.path.exists(path):
            os.remove(path)


# ---------------------------- Cell ----------------------------


@pytest.mark.parametrize(
    "method,operator",
    [
        ("set_rule_greater_than", "greaterThan"),
        ("set_rule_less_than", "lessThan"),
        ("set_rule_equal_to", "equal"),
        ("set_rule_not_equal_to", "notEqual"),
        ("set_rule_greater_than_or_equal_to", "greaterThanOrEqual"),
        ("set_rule_less_than_or_equal_to", "lessThanOrEqual"),
    ],
)
def test_cell_comparison_rules(method, operator):
    cf = ConditionalFormatCell()
    getattr(cf, method)(25)
    rules = _apply(cf)
    assert len(rules) == 1
    assert rules[0].type == "cellIs"
    assert rules[0].operator == operator


def test_cell_between_and_not_between():
    cf = ConditionalFormatCell()
    cf.set_rule_between(15, 35)
    assert _apply(cf)[0].operator == "between"

    cf2 = ConditionalFormatCell()
    cf2.set_rule_not_between(15, 35)
    assert _apply(cf2)[0].operator == "notBetween"


def test_cell_rule_with_format():
    fmt = Format()
    fmt.set_bold()
    fmt.set_background_color("#FFC7CE")
    cf = ConditionalFormatCell()
    cf.set_rule_greater_than(25)
    cf.set_format(fmt)
    rules = _apply(cf)
    assert rules[0].type == "cellIs"
    # A dxf record is what carries the format for a CF rule.
    assert rules[0].dxfId is not None


# ------------------- Blank / Duplicate / Error -------------------


def test_blank_and_inverted_blank():
    assert _apply(ConditionalFormatBlank())[0].type == "containsBlanks"
    cf = ConditionalFormatBlank()
    cf.invert()
    assert _apply(cf)[0].type == "notContainsBlanks"


def test_duplicate_and_inverted_duplicate():
    assert _apply(ConditionalFormatDuplicate())[0].type == "duplicateValues"
    cf = ConditionalFormatDuplicate()
    cf.invert()
    assert _apply(cf)[0].type == "uniqueValues"


def test_error_and_inverted_error():
    assert _apply(ConditionalFormatError())[0].type == "containsErrors"
    cf = ConditionalFormatError()
    cf.invert()
    assert _apply(cf)[0].type == "notContainsErrors"


# ---------------------------- Formula ----------------------------


def test_formula_rule():
    cf = ConditionalFormatFormula()
    cf.set_rule("=$A1>25")
    rules = _apply(cf)
    assert rules[0].type == "expression"


# ---------------------------- Average ----------------------------


@pytest.mark.parametrize(
    "rule", ["above", "below", "equal_or_above", "1_std_dev_above", "3_std_dev_below"]
)
def test_average_rules(rule):
    cf = ConditionalFormatAverage()
    cf.set_rule(rule)
    assert _apply(cf)[0].type == "aboveAverage"


def test_average_rule_invalid_raises():
    cf = ConditionalFormatAverage()
    with pytest.raises(ValueError) as exc:
        cf.set_rule("sideways")
    assert "average rule" in str(exc.value)


# ------------------------------ Top ------------------------------


@pytest.mark.parametrize("kind", ["top", "bottom", "top_percent", "bottom_percent"])
def test_top_rules(kind):
    cf = ConditionalFormatTop()
    cf.set_rule(kind, 3)
    assert _apply(cf)[0].type == "top10"


def test_top_rule_invalid_raises():
    cf = ConditionalFormatTop()
    with pytest.raises(ValueError):
        cf.set_rule("middle", 3)


# ------------------------------ Text ------------------------------


@pytest.mark.parametrize(
    "kind", ["contains", "does_not_contain", "begins_with", "ends_with"]
)
def test_text_rules(kind):
    cf = ConditionalFormatText()
    cf.set_rule(kind, "foo")
    # openpyxl reports the specific text operator as the rule type for
    # begins/ends, and containsText/notContainsText for the other two.
    assert _apply(cf)[0].type in (
        "containsText",
        "notContainsText",
        "beginsWith",
        "endsWith",
    )


def test_text_rule_invalid_raises():
    cf = ConditionalFormatText()
    with pytest.raises(ValueError):
        cf.set_rule("sounds_like", "foo")


# ------------------------------ Date ------------------------------


@pytest.mark.parametrize("rule", ["yesterday", "today", "last_7_days", "next_month"])
def test_date_rules(rule):
    cf = ConditionalFormatDate()
    cf.set_rule(rule)
    assert _apply(cf)[0].type == "timePeriod"


def test_date_rule_invalid_raises():
    cf = ConditionalFormatDate()
    with pytest.raises(ValueError):
        cf.set_rule("next_fortnight")


# --------------------------- Color scales ---------------------------


def test_2_color_scale():
    cf = ConditionalFormat2ColorScale()
    cf.set_minimum_color("#FFFFFF")
    cf.set_maximum_color("#FF0000")
    assert _apply(cf)[0].type == "colorScale"


def test_3_color_scale_with_midpoint():
    cf = ConditionalFormat3ColorScale()
    cf.set_minimum("min", 0)
    cf.set_midpoint("percentile", 50)
    cf.set_maximum("max", 0)
    cf.set_minimum_color("#F8696B")
    cf.set_midpoint_color("#FFEB84")
    cf.set_maximum_color("#63BE7B")
    assert _apply(cf)[0].type == "colorScale"


def test_color_scale_accepts_formula_value_as_string():
    cf = ConditionalFormat2ColorScale()
    cf.set_minimum("formula", "=$A$1")
    cf.set_maximum("max", 0)
    assert _apply(cf)[0].type == "colorScale"


def test_color_scale_invalid_type_raises():
    cf = ConditionalFormat2ColorScale()
    with pytest.raises(ValueError) as exc:
        cf.set_minimum("mediumish", 5)
    assert "conditional format type" in str(exc.value)


def test_color_scale_invalid_value_raises():
    cf = ConditionalFormat2ColorScale()
    with pytest.raises(TypeError):
        cf.set_minimum("number", [1, 2, 3])


# ----------------------------- Data bar -----------------------------


def test_data_bar_basic():
    cf = ConditionalFormatDataBar()
    cf.set_fill_color("#638EC6")
    assert _apply(cf)[0].type == "dataBar"


def test_data_bar_full_option_sweep():
    cf = ConditionalFormatDataBar()
    cf.set_minimum("number", 0)
    cf.set_maximum("number", 100)
    cf.set_fill_color("#638EC6")
    cf.set_border_color("#000000")
    cf.set_negative_fill_color("#FF0000")
    cf.set_negative_border_color("#FF0000")
    cf.set_axis_color("#000000")
    cf.set_solid_fill(True)
    cf.set_border_off(False)
    cf.set_bar_only(True)
    cf.set_direction("left_to_right")
    cf.set_axis_position("midpoint")
    assert _apply(cf)[0].type == "dataBar"


def test_data_bar_classic_style():
    cf = ConditionalFormatDataBar()
    cf.use_classic_style()
    assert _apply(cf)[0].type == "dataBar"


@pytest.mark.parametrize(
    "method,bad", [("set_direction", "sideways"), ("set_axis_position", "diagonal")]
)
def test_data_bar_invalid_enum_raises(method, bad):
    cf = ConditionalFormatDataBar()
    with pytest.raises(ValueError):
        getattr(cf, method)(bad)


def test_data_bar_invalid_color_raises():
    cf = ConditionalFormatDataBar()
    with pytest.raises(ValueError):
        cf.set_fill_color("not-a-color")


# --------------------------- Common methods ---------------------------


def test_multi_range():
    cf = ConditionalFormatCell()
    cf.set_rule_greater_than(25)
    cf.set_multi_range("A1:A5 C1:C5")
    ranges = _ranges(cf)
    joined = " ".join(ranges)
    assert "C1:C5" in joined


def test_stop_if_true():
    cf = ConditionalFormatCell()
    cf.set_rule_greater_than(25)
    cf.set_stop_if_true(True)
    rules = _apply(cf)
    assert rules[0].stopIfTrue


# ------------------------------ Dispatch ------------------------------


def test_add_conditional_format_rejects_non_cf_object():
    wb = Workbook()
    ws = wb.add_worksheet()
    with pytest.raises(TypeError):
        ws.add_conditional_format(0, 0, 4, 0, "not a conditional format")


def test_add_conditional_format_rejects_format_object():
    """A Format is a pyclass too -- the downcast chain must still reject it."""
    wb = Workbook()
    ws = wb.add_worksheet()
    with pytest.raises(TypeError):
        ws.add_conditional_format(0, 0, 4, 0, Format())


def test_same_cf_object_applied_to_two_ranges():
    """Setters clone the inner builder, so reuse must not consume it."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        cf = ConditionalFormatCell()
        cf.set_rule_greater_than(25)
        wb = Workbook()
        ws = wb.add_worksheet()
        ws.write(0, 0, 1)
        ws.add_conditional_format(0, 0, 4, 0, cf)
        ws.add_conditional_format(0, 2, 4, 2, cf)
        wb.close(path)
        sheet = openpyxl.load_workbook(path).active
        rules = [r for c in sheet.conditional_formatting for r in c.rules]
        assert len(rules) == 2
    finally:
        if os.path.exists(path):
            os.remove(path)


def test_two_different_rules_on_one_range():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        high = ConditionalFormatCell()
        high.set_rule_greater_than(40)
        blanks = ConditionalFormatBlank()
        wb = Workbook()
        ws = wb.add_worksheet()
        ws.write(0, 0, 1)
        ws.add_conditional_format(0, 0, 4, 0, high)
        ws.add_conditional_format(0, 0, 4, 0, blanks)
        wb.close(path)
        sheet = openpyxl.load_workbook(path).active
        types = {r.type for c in sheet.conditional_formatting for r in c.rules}
        assert types == {"cellIs", "containsBlanks"}
    finally:
        if os.path.exists(path):
            os.remove(path)
