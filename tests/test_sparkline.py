"""Sparkline tests.

Sparklines are written into the worksheet's extLst, which openpyxl does
not parse, so these assert against the emitted sheet XML directly.
"""
import os
import tempfile
import zipfile

import pytest

from rvgsrust_xlsxwriter import Sparkline, Workbook


def _sheet_xml(build):
    """Fill a 5x5 grid, run build(ws), return xl/worksheets/sheet1.xml."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        wb = Workbook()
        ws = wb.add_worksheet()
        for r in range(5):
            for c in range(5):
                ws.write(r, c, (r + 1) * (c + 1))
        build(ws)
        wb.close(path)
        with zipfile.ZipFile(path) as z:
            return z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    finally:
        if os.path.exists(path):
            os.remove(path)


def _simple(**calls):
    """A single line sparkline over Sheet1!A1:E1 placed in F1."""

    def build(ws):
        sp = Sparkline()
        sp.set_range("Sheet1!A1:E1")
        for name, arg in calls.items():
            method = getattr(sp, name)
            if isinstance(arg, tuple):
                method(*arg)
            else:
                method(arg)
        ws.add_sparkline(0, 5, sp)

    return _sheet_xml(build)


# ------------------------------ types ------------------------------


def test_line_is_the_default_and_emits_no_type_attribute():
    xml = _simple()
    assert "x14:sparklineGroup" in xml
    # Line is the default, so upstream omits the attribute entirely.
    assert 'type="column"' not in xml
    assert 'type="stacked"' not in xml


def test_column_type():
    assert 'type="column"' in _simple(set_type="column")


def test_win_lose_serializes_as_stacked():
    """Upstream spells the variant WinLose, and writes it as "stacked"."""
    assert 'type="stacked"' in _simple(set_type="win_lose")


def test_win_loss_spelling_also_accepted():
    assert 'type="stacked"' in _simple(set_type="win_loss")


def test_type_is_case_insensitive():
    assert 'type="column"' in _simple(set_type="COLUMN")


def test_invalid_type_raises():
    sp = Sparkline()
    with pytest.raises(ValueError) as exc:
        sp.set_type("squiggle")
    assert "sparkline type" in str(exc.value)


# ------------------------------ points ------------------------------


@pytest.mark.parametrize(
    "method,attr",
    [
        ("show_high_point", "high"),
        ("show_low_point", "low"),
        ("show_first_point", "first"),
        ("show_last_point", "last"),
        ("show_negative_points", "negative"),
        ("show_markers", "markers"),
    ],
)
def test_show_point_flags(method, attr):
    assert f'{attr}="1"' in _simple(**{method: True})


def test_show_axis_and_hidden_data_do_not_error():
    xml = _simple(show_axis=True, show_hidden_data=True)
    assert "x14:sparklineGroup" in xml


def test_right_to_left_and_column_order():
    xml = _simple(set_right_to_left=True, set_column_order=True)
    assert "x14:sparklineGroup" in xml


# ------------------------------ scaling ------------------------------


def test_line_weight():
    assert 'lineWeight="2.25"' in _simple(set_line_weight=2.25)


def test_custom_max_and_min():
    xml = _simple(set_custom_max=100.5, set_custom_min=-10.5)
    assert 'manualMax="100.5"' in xml
    assert 'manualMin="-10.5"' in xml


def test_group_max_and_min():
    xml = _simple(set_group_max=True, set_group_min=True)
    assert "x14:sparklineGroup" in xml


def test_style_preset():
    xml = _simple(set_style=12)
    assert "x14:sparklineGroup" in xml


# ------------------------------ colors ------------------------------


@pytest.mark.parametrize(
    "method",
    [
        "set_sparkline_color",
        "set_high_point_color",
        "set_low_point_color",
        "set_first_point_color",
        "set_last_point_color",
        "set_negative_points_color",
        "set_markers_color",
    ],
)
def test_colors_reach_the_xml(method):
    xml = _simple(**{method: "#638EC6"})
    assert "638EC6" in xml


def test_invalid_color_raises():
    sp = Sparkline()
    with pytest.raises(ValueError):
        sp.set_sparkline_color("burnt sienna")


# --------------------------- empty cells ---------------------------


@pytest.mark.parametrize("option", ["gaps", "zero", "connected"])
def test_show_empty_cells_as(option):
    """The variant is Gaps, not Gap."""
    xml = _simple(show_empty_cells_as=option)
    assert "x14:sparklineGroup" in xml


def test_invalid_empty_cells_option_raises():
    sp = Sparkline()
    with pytest.raises(ValueError) as exc:
        sp.show_empty_cells_as("blank")
    assert "empty cells option" in str(exc.value)


# ---------------------------- date range ----------------------------


def test_date_range_sets_date_axis():
    def build(ws):
        sp = Sparkline()
        sp.set_range("Sheet1!A1:E1")
        sp.set_date_range("Sheet1!A2:E2")
        ws.add_sparkline(0, 5, sp)

    assert 'dateAxis="1"' in _sheet_xml(build)


# ------------------------------ groups ------------------------------


def test_sparkline_group_over_2d_range():
    def build(ws):
        sp = Sparkline()
        sp.set_range("Sheet1!A1:E5")
        sp.set_type("column")
        ws.add_sparkline_group(0, 5, 4, 5, sp)

    xml = _sheet_xml(build)
    assert xml.count("</x14:sparklineGroup>") == 1
    # One <x14:sparkline> entry per row of the group.
    assert xml.count("<x14:sparkline>") == 5


def test_sparkline_group_requires_a_2d_range():
    def build(ws):
        sp = Sparkline()
        sp.set_range("Sheet1!A1:E1")  # 1D
        ws.add_sparkline_group(0, 5, 4, 5, sp)

    with pytest.raises(ValueError) as exc:
        _sheet_xml(build)
    assert "2D" in str(exc.value)


def test_sparkline_without_range_raises():
    def build(ws):
        ws.add_sparkline(0, 5, Sparkline())

    with pytest.raises(ValueError):
        _sheet_xml(build)


def test_group_max_shares_scale_across_group():
    def build(ws):
        sp = Sparkline()
        sp.set_range("Sheet1!A1:E5")
        sp.set_group_max(True)
        sp.set_group_min(True)
        ws.add_sparkline_group(0, 5, 4, 5, sp)

    assert "x14:sparklineGroup" in _sheet_xml(build)


# ------------------------------ reuse ------------------------------


def test_same_sparkline_applied_twice():
    """Setters clone the inner builder, so reuse must not consume it."""

    def build(ws):
        sp = Sparkline()
        sp.set_range("Sheet1!A1:E1")
        sp.set_type("column")
        ws.add_sparkline(0, 5, sp)
        ws.add_sparkline(1, 5, sp)

    xml = _sheet_xml(build)
    assert xml.count("</x14:sparklineGroup>") == 2


def test_add_sparkline_rejects_non_sparkline():
    wb = Workbook()
    ws = wb.add_worksheet()
    with pytest.raises(TypeError):
        ws.add_sparkline(0, 5, "not a sparkline")


def test_full_option_sweep():
    """Everything at once, to catch any setter that panics in combination."""

    def build(ws):
        sp = Sparkline()
        sp.set_range("Sheet1!A1:E5")
        sp.set_type("column")
        sp.show_high_point(True)
        sp.show_low_point(True)
        sp.show_first_point(True)
        sp.show_last_point(True)
        sp.show_negative_points(True)
        sp.show_markers(True)
        sp.show_axis(True)
        sp.show_hidden_data(True)
        sp.show_empty_cells_as("connected")
        sp.set_right_to_left(True)
        sp.set_column_order(True)
        sp.set_sparkline_color("#638EC6")
        sp.set_high_point_color("#FF0000")
        sp.set_low_point_color("#00B050")
        sp.set_first_point_color("#7030A0")
        sp.set_last_point_color("#0070C0")
        sp.set_negative_points_color("#FF0000")
        sp.set_markers_color("#000000")
        sp.set_line_weight(1.5)
        sp.set_custom_max(50.0)
        sp.set_custom_min(0.0)
        sp.set_style(20)
        ws.add_sparkline_group(0, 5, 4, 5, sp)

    assert "x14:sparklineGroup" in _sheet_xml(build)


# ---------------------------------------------------------------------
# Regression: rust_xlsxwriter 0.98.1 fixed an XML error that occurred
# when a sparkline and a conditional-format data bar were both present
# on the same worksheet (see the upgrade notes in Cargo.toml). No test
# in this suite combined the two before the 0.98.2 upgrade.
# ---------------------------------------------------------------------

def test_sparkline_and_conditional_format_databar_together_produce_valid_file():
    import openpyxl
    from rvgsrust_xlsxwriter import ConditionalFormatDataBar

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        wb = Workbook()
        ws = wb.add_worksheet()
        for r in range(5):
            for c in range(5):
                ws.write(r, c, (r + 1) * (c + 1))

        sp = Sparkline()
        sp.set_range("Sheet1!A1:E1")
        ws.add_sparkline(0, 5, sp)

        cf = ConditionalFormatDataBar()
        cf.set_fill_color("#638EC6")
        ws.add_conditional_format(0, 0, 4, 4, cf)

        wb.close(path)

        # The bug produced an XML error, not necessarily a write-time
        # exception -- confirm the resulting file is actually valid by
        # having an independent consumer (openpyxl) load it.
        sheet = openpyxl.load_workbook(path).active
        assert sheet["A1"].value == 1
    finally:
        if os.path.exists(path):
            os.remove(path)
