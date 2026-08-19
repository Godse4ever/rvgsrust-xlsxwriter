"""Column, row, cell and range format tests.

Cell-level formats are checked through openpyxl. Column and row formats
live as style attributes on the <col> and <row> elements rather than on
individual cells, so those are checked against the sheet XML.
"""
import os
import re
import tempfile
import zipfile

import openpyxl
import pytest

from rvgsrust_xlsxwriter import Format, Workbook


def _build(build):
    """Run build(ws), return (openpyxl sheet, sheet1.xml, styles.xml)."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        wb = Workbook()
        ws = wb.add_worksheet()
        build(ws)
        wb.close(path)
        with zipfile.ZipFile(path) as z:
            sheet_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
            styles_xml = z.read("xl/styles.xml").decode("utf-8")
        return openpyxl.load_workbook(path).active, sheet_xml, styles_xml
    finally:
        if os.path.exists(path):
            os.remove(path)


def _num_format():
    fmt = Format()
    fmt.set_num_format("0.00")
    return fmt


# ---------------------------- column ----------------------------


def test_set_column_format():
    def build(ws):
        ws.write(0, 1, 1.5)
        ws.set_column_format(1, _num_format())

    _, sheet_xml, styles_xml = _build(build)
    assert re.search(r'<col[^>]*style="\d+"', sheet_xml), sheet_xml
    assert "0.00" in styles_xml


def test_set_column_range_format():
    def build(ws):
        ws.set_column_range_format(1, 3, _num_format())

    _, sheet_xml, styles_xml = _build(build)
    cols = re.findall(r'<col[^>]*style="\d+"[^>]*>', sheet_xml)
    assert cols, sheet_xml
    assert "0.00" in styles_xml


def test_set_column_range_format_rejects_reversed_range():
    def build(ws):
        ws.set_column_range_format(5, 2, _num_format())

    with pytest.raises(ValueError):
        _build(build)


# ------------------------------ row ------------------------------


def test_set_row_format():
    def build(ws):
        ws.write(2, 0, 1.5)
        ws.set_row_format(2, _num_format())

    _, sheet_xml, styles_xml = _build(build)
    assert re.search(r'<row[^>]*s="\d+"', sheet_xml), sheet_xml
    assert "0.00" in styles_xml


# --------------------------- cell / range ---------------------------


def test_set_cell_format():
    def build(ws):
        ws.write(0, 0, 1.5)
        ws.set_cell_format(0, 0, _num_format())

    sheet, _, _ = _build(build)
    assert sheet["A1"].number_format == "0.00"


def test_set_cell_format_on_an_empty_cell():
    def build(ws):
        ws.set_cell_format(0, 0, _num_format())

    sheet, _, _ = _build(build)
    assert sheet["A1"].number_format == "0.00"


def test_set_range_format():
    def build(ws):
        for r in range(3):
            for c in range(2):
                ws.write(r, c, r + c)
        ws.set_range_format(0, 0, 2, 1, _num_format())

    sheet, _, _ = _build(build)
    for ref in ("A1", "B1", "A3", "B3"):
        assert sheet[ref].number_format == "0.00", ref


def test_set_range_format_preserves_values():
    def build(ws):
        ws.write(0, 0, 42)
        ws.set_range_format(0, 0, 1, 1, _num_format())

    sheet, _, _ = _build(build)
    assert sheet["A1"].value == 42


def test_set_range_format_with_a_date_format():
    fmt = Format()
    fmt.set_num_format("yyyy-mm-dd")

    def build(ws):
        ws.write(0, 0, 45123)
        ws.set_range_format(0, 0, 0, 0, fmt)

    sheet, _, _ = _build(build)
    assert sheet["A1"].number_format == "yyyy-mm-dd"


# --------------------------- motivating case ---------------------------


def test_reformatting_a_dataframe_column_after_the_fact():
    """The gap this closes: before set_column_format there was no way to
    change a column's format after write_dataframe had written it."""
    pd = pytest.importorskip("pandas")

    def build(ws):
        df = pd.DataFrame({"amount": [1.5, 2.5, 3.5]})
        ws.write_dataframe(0, 0, df)
        fmt = Format()
        fmt.set_num_format("#,##0.00")
        ws.set_column_format(0, fmt)

    _, sheet_xml, styles_xml = _build(build)
    assert re.search(r'<col[^>]*style="\d+"', sheet_xml)
    assert "#,##0.00" in styles_xml


# ------------------------- constant memory -------------------------


def test_set_row_format_backward_write_raises_in_constant_memory():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        wb = Workbook()
        ws = wb.add_worksheet(constant_memory=True)
        ws.write(5, 0, 1)
        with pytest.raises(ValueError):
            ws.set_row_format(2, _num_format())
    finally:
        if os.path.exists(path):
            os.remove(path)


def test_set_cell_format_backward_write_raises_in_constant_memory():
    wb = Workbook()
    ws = wb.add_worksheet(constant_memory=True)
    ws.write(5, 0, 1)
    with pytest.raises(ValueError):
        ws.set_cell_format(2, 0, _num_format())


def test_column_format_unaffected_by_constant_memory_row_order():
    """A column format targets no particular row, so it is not guarded."""

    def build(ws):
        ws.write(3, 0, 1)
        ws.set_column_format(0, _num_format())

    _, sheet_xml, _ = _build(build)
    assert re.search(r'<col[^>]*style="\d+"', sheet_xml)


# ---------------------------- type checks ----------------------------


def test_format_methods_reject_non_format():
    wb = Workbook()
    ws = wb.add_worksheet()
    for call in (
        lambda: ws.set_column_format(0, "not a format"),
        lambda: ws.set_row_format(0, "not a format"),
        lambda: ws.set_cell_format(0, 0, "not a format"),
        lambda: ws.set_range_format(0, 0, 1, 1, "not a format"),
    ):
        with pytest.raises(TypeError):
            call()


# ------------------- set_range_format_with_border -------------------


def test_range_format_with_border_applies_cell_and_border_formats():
    def build(ws):
        cell_fmt = Format()
        cell_fmt.set_background_color("#FFFF00")
        border_fmt = Format()
        border_fmt.set_border("thin")
        for r in range(3):
            for c in range(3):
                ws.write(r, c, r * 3 + c)
        ws.set_range_format_with_border(0, 0, 2, 2, cell_fmt, border_fmt)

    sheet, _, styles = _build(build)
    assert sheet["A1"].value == 0
    assert sheet["C3"].value == 8
    # Both the fill (cell_format) and a border (border_format) should
    # show up somewhere in styles.xml -- rust_xlsxwriter builds up to
    # 9 distinct per-position style combinations internally, so this
    # doesn't assert an exact count, just that both properties landed.
    assert "FFFF00" in styles
    assert "<left" in styles or "<right" in styles or "<top" in styles


def test_range_format_with_border_rejects_reversed_range():
    def build(ws):
        cell_fmt = Format()
        border_fmt = Format()
        border_fmt.set_border("thin")
        ws.set_range_format_with_border(5, 5, 0, 0, cell_fmt, border_fmt)

    with pytest.raises(ValueError):
        _build(build)


def test_range_format_with_border_rejects_non_format():
    wb = Workbook()
    ws = wb.add_worksheet()
    valid_fmt = Format()
    with pytest.raises(TypeError):
        ws.set_range_format_with_border(0, 0, 1, 1, "not a format", valid_fmt)
    with pytest.raises(TypeError):
        ws.set_range_format_with_border(0, 0, 1, 1, valid_fmt, "not a format")


# ---------------------------- clear_cell_format ----------------------------


def test_clear_cell_format_removes_formatting_keeps_value():
    def build(ws):
        fmt = Format()
        fmt.set_bold()
        ws.write(0, 0, "styled", fmt)
        ws.write(0, 1, "also styled", fmt)
        ws.clear_cell_format(0, 0)

    sheet, _, _ = _build(build)
    # Value survives the format clear.
    assert sheet["A1"].value == "styled"
    assert sheet["B1"].value == "also styled"
    # A1's bold is gone; B1's remains, proving clear_cell_format()
    # targeted only the specified cell.
    assert not sheet["A1"].font.bold
    assert sheet["B1"].font.bold


def test_clear_cell_format_on_unformatted_cell_is_noop():
    def build(ws):
        ws.write(0, 0, "plain")
        ws.clear_cell_format(0, 0)  # must not raise

    sheet, _, _ = _build(build)
    assert sheet["A1"].value == "plain"


def test_clear_cell_format_on_nonexistent_cell_is_noop():
    def build(ws):
        ws.write(0, 0, "x")
        ws.clear_cell_format(50, 50)  # never written, must not raise

    _build(build)
