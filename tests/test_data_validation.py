"""Data validation tests.

Verified through openpyxl's DataValidation reader rather than raw XML,
matching this project's convention for conditional formats.
"""
import os
import tempfile

import openpyxl
import pytest

from rvgsrust_xlsxwriter import DataValidation, Workbook


def _apply(build, rows=5):
    """Write a small numeric column, run build(ws, dv), reload via openpyxl.
    Returns (sheet, list of openpyxl DataValidation objects)."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        wb = Workbook()
        ws = wb.add_worksheet()
        for i in range(rows):
            ws.write(i, 0, i)
        build(ws)
        wb.close(path)
        sheet = openpyxl.load_workbook(path).active
        return sheet, list(sheet.data_validations.dataValidation)
    finally:
        if os.path.exists(path):
            os.remove(path)


# ---------------------------- whole number ----------------------------


def test_whole_number_between():
    def build(ws):
        dv = DataValidation()
        dv.allow_whole_number("between", 1, 10)
        ws.add_data_validation(0, 0, 4, 0, dv)

    _, dvs = _apply(build)
    assert len(dvs) == 1
    d = dvs[0]
    assert d.type == "whole"
    # "between" is the OOXML schema default, so rust_xlsxwriter omits
    # the operator attribute entirely for it (confirmed against
    # worksheet.rs write_data_validation) -- only NotBetween and the
    # 6 single-value comparisons get an explicit operator written.
    # Not asserting on d.operator here for that reason.
    assert d.formula1 == "1"
    assert d.formula2 == "10"
    assert "A1:A5" in str(d.sqref)


def test_whole_number_greater_than():
    def build(ws):
        dv = DataValidation()
        dv.allow_whole_number("greater_than", 5)
        ws.add_data_validation(0, 0, 4, 0, dv)

    _, dvs = _apply(build)
    d = dvs[0]
    assert d.type == "whole"
    assert d.operator == "greaterThan"
    assert d.formula1 == "5"


def test_whole_number_between_missing_second_value_raises():
    dv = DataValidation()
    with pytest.raises(ValueError):
        dv.allow_whole_number("between", 1)


def test_whole_number_invalid_rule_type_raises():
    dv = DataValidation()
    with pytest.raises(ValueError):
        dv.allow_whole_number("close_to", 1)


# ---------------------------- decimal number ----------------------------


def test_decimal_number_not_between():
    def build(ws):
        dv = DataValidation()
        dv.allow_decimal_number("not_between", 0.0, 1.0)
        ws.add_data_validation(0, 0, 4, 0, dv)

    _, dvs = _apply(build)
    d = dvs[0]
    assert d.type == "decimal"
    assert d.operator == "notBetween"


# ---------------------------- text length ----------------------------


def test_text_length_less_than_or_equal_to():
    def build(ws):
        dv = DataValidation()
        dv.allow_text_length("less_than_or_equal_to", 50)
        ws.add_data_validation(0, 0, 4, 0, dv)

    _, dvs = _apply(build)
    d = dvs[0]
    assert d.type == "textLength"
    assert d.operator == "lessThanOrEqual"
    assert d.formula1 == "50"


# ---------------------------- lists (the main feature) ----------------------------


def test_allow_list_strings_creates_dropdown():
    def build(ws):
        dv = DataValidation()
        dv.allow_list_strings(["Yes", "No", "Maybe"])
        ws.add_data_validation(0, 0, 4, 0, dv)

    _, dvs = _apply(build)
    assert len(dvs) == 1
    d = dvs[0]
    assert d.type == "list"
    assert "Yes" in d.formula1
    assert "No" in d.formula1
    assert "Maybe" in d.formula1


def test_allow_list_strings_over_255_chars_raises():
    dv = DataValidation()
    long_list = [f"option_number_{i}" for i in range(30)]  # comfortably > 255 chars joined
    with pytest.raises(ValueError):
        dv.allow_list_strings(long_list)


def test_allow_list_formula_creates_dropdown_from_range():
    def build(ws):
        ws.write_column(0, 5, ["Pass", "Fail", "Incomplete"])
        dv = DataValidation()
        dv.allow_list_formula("F1:F3")
        ws.add_data_validation(1, 3, 1, 3, dv)

    _, dvs = _apply(build)
    d = dvs[0]
    assert d.type == "list"
    assert "F1" in d.formula1 and "F3" in d.formula1


# ---------------------------- custom / any value ----------------------------


def test_allow_custom_formula():
    def build(ws):
        dv = DataValidation()
        dv.allow_custom("=A1>0")
        ws.add_data_validation(0, 0, 4, 0, dv)

    _, dvs = _apply(build)
    d = dvs[0]
    assert d.type == "custom"
    # Formula::new() strips a leading '=' before storing, so check
    # substring rather than exact match -- not asserting on whether
    # the '=' survives round-trip.
    assert "A1>0" in d.formula1


def test_allow_any_value_after_rule_clears_it():
    def build(ws):
        dv = DataValidation()
        dv.allow_whole_number("greater_than", 5)
        dv.allow_any_value()
        dv.show_input_message(True)
        dv.set_input_title("Note")
        ws.add_data_validation(0, 0, 4, 0, dv)

    _, dvs = _apply(build)
    d = dvs[0]
    # "none" is openpyxl's representation of Excel's default/unrestricted type
    assert d.type in ("none", None, "")
    assert d.promptTitle == "Note"


# ---------------------------- messages, error style, multi-range ----------------------------


def test_input_and_error_messages():
    def build(ws):
        dv = DataValidation()
        dv.allow_whole_number("between", 1, 5)
        dv.show_input_message(True)
        dv.set_input_title("Enter a rating")
        dv.set_input_message("1 to 5 only")
        dv.show_error_message(True)
        dv.set_error_title("Invalid rating")
        dv.set_error_message("Must be 1-5")
        dv.set_error_style("warning")
        ws.add_data_validation(0, 0, 4, 0, dv)

    _, dvs = _apply(build)
    d = dvs[0]
    assert d.promptTitle == "Enter a rating"
    assert d.prompt == "1 to 5 only"
    assert d.errorTitle == "Invalid rating"
    assert d.error == "Must be 1-5"
    assert d.errorStyle == "warning"


def test_input_title_over_32_chars_raises():
    dv = DataValidation()
    with pytest.raises(ValueError):
        dv.set_input_title("x" * 40)


def test_error_message_over_255_chars_raises():
    dv = DataValidation()
    with pytest.raises(ValueError):
        dv.set_error_message("x" * 300)


def test_set_error_style_invalid_raises():
    dv = DataValidation()
    with pytest.raises(ValueError):
        dv.set_error_style("critical")


def test_multi_range():
    def build(ws):
        for i in range(5):
            ws.write(i, 2, i)  # column C too
        dv = DataValidation()
        dv.allow_whole_number("greater_than", 0)
        dv.set_multi_range("C1:C5")
        ws.add_data_validation(0, 0, 4, 0, dv)

    _, dvs = _apply(build)
    d = dvs[0]
    sqref = str(d.sqref)
    assert "A1:A5" in sqref
    assert "C1:C5" in sqref


# ---------------------------- constant_memory ----------------------------


def test_data_validation_works_in_constant_memory_mode():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        wb = Workbook()
        ws = wb.add_worksheet(constant_memory=True)
        for i in range(5):
            ws.write(i, 0, i)
        dv = DataValidation()
        dv.allow_whole_number("between", 0, 10)
        ws.add_data_validation(0, 0, 4, 0, dv)
        wb.close(path)
        sheet = openpyxl.load_workbook(path).active
        dvs = list(sheet.data_validations.dataValidation)
        assert len(dvs) == 1
        assert dvs[0].type == "whole"
    finally:
        if os.path.exists(path):
            os.remove(path)
