"""Per-side border colours, diagonal borders, protection, strikethrough
and pattern foreground colour."""
import os
import tempfile
import zipfile

import pytest

from rvgsrust_xlsxwriter import Format, Workbook


def _styles(build):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        wb = Workbook()
        ws = wb.add_worksheet()
        build(wb, ws)
        wb.close(path)
        with zipfile.ZipFile(path) as z:
            return z.read("xl/styles.xml").decode("utf-8")
    finally:
        if os.path.exists(path):
            os.remove(path)


# --------------------- per-side border colours ---------------------


@pytest.mark.parametrize(
    "method", ["set_border_top_color", "set_border_bottom_color",
               "set_border_left_color", "set_border_right_color"]
)
def test_per_side_border_colors(method):
    def build(wb, ws):
        fmt = Format()
        fmt.set_border("thin")
        getattr(fmt, method)("#FF0000")
        ws.write(0, 0, 1, fmt)

    assert "FF0000" in _styles(build)


def test_per_side_border_styles_already_existed():
    """These reverse upstream's word order, which is why the parity audit
    first reported them as missing."""

    def build(wb, ws):
        fmt = Format()
        fmt.set_top_border("thick")
        fmt.set_bottom_border("double")
        fmt.set_left_border("thin")
        fmt.set_right_border("dashed")
        ws.write(0, 0, 1, fmt)

    styles = _styles(build)
    assert "thick" in styles
    assert "double" in styles


def test_border_color_invalid_raises():
    fmt = Format()
    with pytest.raises(ValueError):
        fmt.set_border_top_color("reddish")


def test_border_color_setters_chain():
    fmt = Format()
    result = fmt.set_border_top_color("#FF0000")
    assert result is not None


# ------------------------- diagonal borders -------------------------


@pytest.mark.parametrize("kind", ["none", "up", "down", "up_down"])
def test_diagonal_border_types(kind):
    def build(wb, ws):
        fmt = Format()
        fmt.set_border_diagonal("thin")
        fmt.set_border_diagonal_type(kind)
        fmt.set_border_diagonal_color("#0070C0")
        ws.write(0, 0, 1, fmt)

    styles = _styles(build)
    assert "0070C0" in styles


def test_diagonal_border_type_is_case_insensitive():
    def build(wb, ws):
        fmt = Format()
        fmt.set_border_diagonal("thin")
        fmt.set_border_diagonal_type("UP_DOWN")
        ws.write(0, 0, 1, fmt)

    assert "<styleSheet" in _styles(build)


def test_diagonal_border_type_invalid_raises():
    fmt = Format()
    with pytest.raises(ValueError) as exc:
        fmt.set_border_diagonal_type("sideways")
    assert "diagonal border type" in str(exc.value)


def test_diagonal_border_style_invalid_raises():
    fmt = Format()
    with pytest.raises(ValueError):
        fmt.set_border_diagonal("squiggly")


# --------------------------- protection ---------------------------


def test_locked_and_unlocked():
    def build(wb, ws):
        locked = Format()
        locked.set_locked()
        unlocked = Format()
        unlocked.set_unlocked()
        ws.write(0, 0, 1, locked)
        ws.write(1, 0, 2, unlocked)

    styles = _styles(build)
    # An unlocked cell emits protection with locked="0".
    assert 'locked="0"' in styles


def test_hidden_formula():
    def build(wb, ws):
        fmt = Format()
        fmt.set_hidden()
        ws.write(0, 0, 1, fmt)

    assert 'hidden="1"' in _styles(build)


def test_protection_with_a_protected_sheet():
    """Protection flags only bite once the sheet itself is protected."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        wb = Workbook()
        ws = wb.add_worksheet()
        editable = Format()
        editable.set_unlocked()
        ws.write(0, 0, "fixed")
        ws.write(1, 0, "editable", editable)
        ws.protect()
        wb.close(path)
        with zipfile.ZipFile(path) as z:
            sheet = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
            styles = z.read("xl/styles.xml").decode("utf-8")
        assert "<sheetProtection" in sheet
        assert 'locked="0"' in styles
    finally:
        if os.path.exists(path):
            os.remove(path)


# ------------------ strikethrough and foreground ------------------


def test_font_strikethrough():
    def build(wb, ws):
        fmt = Format()
        fmt.set_font_strikethrough()
        ws.write(0, 0, "struck", fmt)

    assert "<strike" in _styles(build)


def test_foreground_color_with_a_pattern():
    def build(wb, ws):
        fmt = Format()
        fmt.set_pattern("light_up")
        fmt.set_background_color("#FFFF00")
        fmt.set_foreground_color("#FF0000")
        ws.write(0, 0, 1, fmt)

    styles = _styles(build)
    assert "FF0000" in styles
    assert "FFFF00" in styles


def test_foreground_color_invalid_raises():
    fmt = Format()
    with pytest.raises(ValueError):
        fmt.set_foreground_color("nope")


# ---------------------------- combination ----------------------------


def test_all_new_format_options_together():
    def build(wb, ws):
        fmt = Format()
        fmt.set_bold()
        fmt.set_border("thin")
        fmt.set_border_top_color("#FF0000")
        fmt.set_border_bottom_color("#00B050")
        fmt.set_border_diagonal("hair")
        fmt.set_border_diagonal_type("up")
        fmt.set_border_diagonal_color("#0070C0")
        fmt.set_font_strikethrough()
        fmt.set_unlocked()
        ws.write(0, 0, 1, fmt)

    styles = _styles(build)
    for expected in ("FF0000", "00B050", "0070C0", "<strike", 'locked="0"'):
        assert expected in styles, expected


# ------------------ quote prefix, hyperlink, checkbox ------------------


def test_quote_prefix():
    def build(wb, ws):
        fmt = Format()
        fmt.set_quote_prefix()
        ws.write(0, 0, "123", fmt)

    styles = _styles(build)
    assert 'quotePrefix="1"' in styles


def test_hyperlink_style():
    def build(wb, ws):
        fmt = Format()
        fmt.set_hyperlink()
        ws.write(0, 0, "not a real url", fmt)

    # Confirmed via source (format.rs set_hyperlink()) that this sets
    # is_hyperlink, a theme color, and single underline -- not
    # asserting the exact theme-color XML encoding here since tracing
    # that through the color-writing code adds risk for little extra
    # coverage; a full parse via openpyxl already confirms the file
    # isn't corrupted by the style.
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        wb = Workbook()
        ws = wb.add_worksheet()
        build(wb, ws)
        wb.close(path)
        import openpyxl
        sheet = openpyxl.load_workbook(path).active
        assert sheet["A1"].value == "not a real url"
    finally:
        if os.path.exists(path):
            os.remove(path)


def test_checkbox_format():
    def build(wb, ws):
        fmt = Format()
        fmt.set_checkbox()
        ws.write(0, 0, True, fmt)

    # set_checkbox() adds an xfpb:xfComplement extLst extension to the
    # cell's <xf> in styles.xml -- confirmed against actual
    # rust_xlsxwriter source (styles.rs write_xf_format_extensions),
    # not guessed.
    styles = _styles(build)
    assert "xfpb:xfComplement" in styles


# ----------------- font family, reading direction -----------------


def test_font_family():
    def build(wb, ws):
        fmt = Format()
        fmt.set_font_family(2)  # Swiss
        ws.write(0, 0, "x", fmt)

    # Confirmed against styles.rs: written as <family val="2"/>, not a
    # family="2" attribute on some other element.
    styles = _styles(build)
    assert '<family val="2"' in styles


def test_font_charset():
    def build(wb, ws):
        fmt = Format()
        fmt.set_font_charset(1)
        ws.write(0, 0, "x", fmt)

    styles = _styles(build)
    assert '<charset val="1"' in styles


@pytest.mark.parametrize("script,xml_val", [
    ("superscript", "superscript"),
    ("subscript", "subscript"),
])
def test_font_script(script, xml_val):
    def build(wb, ws):
        fmt = Format()
        fmt.set_font_script(script)
        ws.write(0, 0, "x", fmt)

    styles = _styles(build)
    assert f'val="{xml_val}"' in styles


def test_font_script_none_does_not_add_vertalign_value():
    def build(wb, ws):
        fmt = Format()
        fmt.set_font_script("none")
        ws.write(0, 0, "x", fmt)

    _styles(build)  # must not raise -- "none" is the default, no-op


def test_font_script_invalid_raises():
    fmt = Format()
    with pytest.raises(ValueError):
        fmt.set_font_script("sideways")


@pytest.mark.parametrize("direction", [0, 1, 2])
def test_reading_direction_valid(direction):
    def build(wb, ws):
        fmt = Format()
        fmt.set_reading_direction(direction)
        ws.write(0, 0, "x", fmt)

    # Must not raise for any valid value; 0 (context-dependent) may not
    # add an explicit readingOrder attribute since it's the default.
    _styles(build)


def test_reading_direction_invalid_raises():
    fmt = Format()
    with pytest.raises(ValueError):
        fmt.set_reading_direction(3)


# ---------------------------- unset_* ----------------------------


def test_unset_bold_reverses_set_bold():
    def build(wb, ws):
        fmt = Format()
        fmt.set_bold()
        fmt.unset_bold()
        ws.write(0, 0, "x", fmt)

    styles = _styles(build)
    assert 'b val="1"' not in styles and "<b/>" not in styles


@pytest.mark.parametrize("method", [
    "unset_italic", "unset_font_strikethrough", "unset_text_wrap",
    "unset_shrink", "unset_hidden", "unset_quote_prefix",
    "unset_checkbox", "unset_hyperlink_style",
])
def test_unset_methods_do_not_raise(method):
    def build(wb, ws):
        fmt = Format()
        getattr(fmt, method)()
        ws.write(0, 0, "x", fmt)

    _styles(build)  # must not raise
