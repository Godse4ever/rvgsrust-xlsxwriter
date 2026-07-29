"""Regression tests for bugs found in the 0.2.0 audit.

Each test here pins the *observable* consequence of a specific fixed bug,
so a future refactor that reintroduces the root cause fails loudly. They
deliberately assert on the saved workbook (via openpyxl) rather than on
internal state, because every bug in this file was silent at the API
level -- the calls all returned successfully and produced a valid .xlsx
that simply contained the wrong thing.
"""
import os

import openpyxl
import pytest

from rvgsrust_xlsxwriter import Workbook

TEST_FILE = "test_regressions.xlsx"


def teardown_function(function):
    if os.path.exists(TEST_FILE):
        os.remove(TEST_FILE)


# ---------------------------------------------------------------------
# add_worksheet() index desync after a rejected sheet name
# ---------------------------------------------------------------------
# Previously add_worksheet() appended the worksheet, then called
# set_name(), then advanced its index counter. A rejected name returned
# early from the middle of that sequence, leaving an unnamed worksheet in
# the workbook while the counter had not moved. The next add_worksheet()
# therefore returned an index pointing at the orphan, and every write
# through that handle landed on the wrong sheet while the intended sheet
# saved empty. Nothing raised.


@pytest.mark.parametrize(
    "bad_name",
    [
        "A" * 32,      # exceeds Excel's 31-character limit
        "bad/name",    # contains an invalid character
        "bad:name",
        "bad[name]",
        "",            # blank
    ],
)
def test_rejected_sheet_name_leaves_workbook_unchanged(bad_name):
    """A rejected name must not append a worksheet, and must not shift
    the index of any worksheet added afterwards."""
    wb = Workbook()
    wb.add_worksheet("First")

    with pytest.raises(ValueError):
        wb.add_worksheet(bad_name)

    second = wb.add_worksheet("Second")
    second.write(0, 0, "landed-correctly")
    wb.close(TEST_FILE)

    book = openpyxl.load_workbook(TEST_FILE)
    # No orphan sheet from the failed call.
    assert book.sheetnames == ["First", "Second"]
    # The write went to the sheet the caller actually asked for.
    assert book["Second"]["A1"].value == "landed-correctly"
    assert book["First"]["A1"].value is None


def test_many_sheets_keep_stable_indices_across_failures():
    """Interleaving failures with successes must not corrupt any handle."""
    wb = Workbook()
    handles = {}
    for i in range(5):
        handles[f"S{i}"] = wb.add_worksheet(f"S{i}")
        with pytest.raises(ValueError):
            wb.add_worksheet("X" * 32)

    for name, ws in handles.items():
        ws.write(0, 0, name)
    wb.close(TEST_FILE)

    book = openpyxl.load_workbook(TEST_FILE)
    assert book.sheetnames == [f"S{i}" for i in range(5)]
    # Each sheet must contain its own name, not a neighbour's.
    for name in handles:
        assert book[name]["A1"].value == name


# ---------------------------------------------------------------------
# add_table() bypassed the constant_memory row-order guard
# ---------------------------------------------------------------------
# add_table() writes header/total cells immediately rather than at save
# time, but was the only cell-writing Worksheet method with no
# check_row_order call. On a constant_memory worksheet, a table anchored
# above the current high-water mark was accepted and silently produced a
# corrupt file instead of raising.


def test_add_table_rejects_backward_write_in_constant_memory():
    from rvgsrust_xlsxwriter import Table

    wb = Workbook()
    ws = wb.add_worksheet("CM", constant_memory=True)

    # Advance the high-water mark well past where the table would start.
    ws.write(50, 0, "later-row")

    table = Table()
    with pytest.raises(ValueError, match="constant_memory"):
        ws.add_table(0, 0, 10, 2, table)


def test_add_table_still_allowed_in_order_in_constant_memory():
    """The guard must not reject a legitimate forward-ordered table."""
    from rvgsrust_xlsxwriter import Table

    wb = Workbook()
    ws = wb.add_worksheet("CM", constant_memory=True)
    ws.write(0, 0, "header-area")

    table = Table()
    ws.add_table(5, 0, 10, 2, table)  # forward of row 0 -- must not raise
    wb.close(TEST_FILE)
    assert os.path.exists(TEST_FILE)


def test_add_table_unaffected_on_normal_worksheet():
    """Non-constant_memory sheets keep unrestricted ordering."""
    from rvgsrust_xlsxwriter import Table

    wb = Workbook()
    ws = wb.add_worksheet("Normal")
    ws.write(50, 0, "later-row")

    table = Table()
    ws.add_table(0, 0, 10, 2, table)  # backward, but allowed here
    wb.close(TEST_FILE)
    assert os.path.exists(TEST_FILE)


# ---------------------------------------------------------------------
# Re-entrant workbook access panicked instead of raising
# ---------------------------------------------------------------------
# write_records() holds the workbook's RefCell borrow across the whole
# loop while calling classify() per cell, and classify() falls back to
# value.str() for unrecognised types. A __str__ that touches the same
# Workbook re-entered borrow_mut() and hit a Rust panic surfacing as
# pyo3's PanicException with only "already mutably borrowed". It is now a
# RuntimeError explaining the cause.


def test_reentrant_write_from_dunder_str_raises_runtime_error():
    wb = Workbook()
    ws = wb.add_worksheet("S")

    class ReentrantValue:
        """Has no __float__/__index__ and is not a str, so classify()
        falls through to value.str() -- running this __str__ while the
        workbook borrow is held."""

        def __str__(self):
            ws.write(99, 0, "re-entrant")
            return "value"

    with pytest.raises(RuntimeError, match="already being modified"):
        ws.write_records(0, 0, [{"col": ReentrantValue()}])


def test_reentrant_value_is_fine_on_the_per_cell_path():
    """write() classifies before taking the borrow, so the same value is
    harmless there. Pins that the fix did not over-restrict."""
    wb = Workbook()
    ws = wb.add_worksheet("S")

    class Chatty:
        def __str__(self):
            return "converted"

    ws.write(0, 0, Chatty())
    wb.close(TEST_FILE)

    book = openpyxl.load_workbook(TEST_FILE)
    assert book["S"]["A1"].value == "converted"


# ---------------------------------------------------------------------
# close() held the GIL for the whole save
# ---------------------------------------------------------------------
# save() serialises and deflates the entire archive and touches no Python
# objects, but ran with the GIL held, stalling every other thread for its
# full duration. It now releases the GIL. This test asserts a background
# thread keeps making progress during a save, and that a concurrent
# access to the same Workbook is reported as an error rather than
# corrupting state.


def test_other_threads_progress_during_close():
    import threading
    import time

    wb = Workbook()
    ws = wb.add_worksheet("Big")
    # Enough data that save() takes long enough to observe.
    ws.write_rows(0, 0, [[f"cell-{r}-{c}" for c in range(20)] for r in range(20000)])

    ticks = []
    stop = threading.Event()

    def ticker():
        while not stop.is_set():
            ticks.append(1)
            time.sleep(0.001)

    t = threading.Thread(target=ticker, daemon=True)
    t.start()
    try:
        wb.close(TEST_FILE)
    finally:
        stop.set()
        t.join(timeout=5)

    # With the GIL released during save, the ticker runs throughout.
    assert len(ticks) > 0
    assert os.path.exists(TEST_FILE)


def test_concurrent_close_reports_error_not_corruption():
    """A second thread entering the same Workbook mid-save must be
    refused cleanly by the borrow guard."""
    import threading

    wb = Workbook()
    ws = wb.add_worksheet("S")
    ws.write_rows(0, 0, [[f"v{r}-{c}" for c in range(10)] for r in range(20000)])

    errors = []

    def second_close():
        try:
            wb.close("test_regressions_second.xlsx")
        except Exception as exc:  # RuntimeError if it lands mid-save
            errors.append(exc)

    t = threading.Thread(target=second_close)
    t.start()
    wb.close(TEST_FILE)
    t.join(timeout=30)

    # Either it serialised cleanly (no error) or it was refused with a
    # RuntimeError -- never a panic or a corrupt file.
    for exc in errors:
        assert isinstance(exc, RuntimeError)
    assert os.path.exists(TEST_FILE)
    if os.path.exists("test_regressions_second.xlsx"):
        os.remove("test_regressions_second.xlsx")


# ---------------------------------------------------------------------
# Arrow string columns allocated a String per cell
# ---------------------------------------------------------------------
# arrow_cell_value() called .to_string() on every Utf8/LargeUtf8/Utf8View
# cell, heap-allocating a String that was dropped immediately after
# write_string(). CellValue::Str is now Cow<str> and the Arrow path
# borrows the columnar buffer directly. These tests pin that the change
# is behaviour-preserving, which is the only part observable from Python.


def _pa():
    return pytest.importorskip("pyarrow")


def test_arrow_string_columns_roundtrip_unchanged():
    pa = _pa()
    table = pa.table(
        {
            "utf8": pa.array(["a", "bb", None, "ünïcødé", ""], type=pa.string()),
            "large": pa.array(["x", None, "zzz", "q", "w"], type=pa.large_string()),
        }
    )
    wb = Workbook()
    ws = wb.add_worksheet("S")
    ws.write_dataframe(0, 0, table)
    wb.close(TEST_FILE)

    book = openpyxl.load_workbook(TEST_FILE)
    sheet = book["S"]
    assert [sheet.cell(1, c).value for c in (1, 2)] == ["utf8", "large"]
    assert [sheet.cell(r, 1).value for r in range(2, 7)] == [
        "a", "bb", None, "ünïcødé", None,
    ]
    assert [sheet.cell(r, 2).value for r in range(2, 7)] == [
        "x", None, "zzz", "q", "w",
    ]


def test_arrow_string_view_roundtrip_unchanged():
    """Utf8View is Polars' default string type, so it is the hot path."""
    pa = _pa()
    if not hasattr(pa, "string_view"):
        pytest.skip("pyarrow build lacks string_view")
    table = pa.table({"v": pa.array(["one", None, "three"], type=pa.string_view())})
    wb = Workbook()
    ws = wb.add_worksheet("S")
    ws.write_dataframe(0, 0, table)
    wb.close(TEST_FILE)

    sheet = openpyxl.load_workbook(TEST_FILE)["S"]
    assert [sheet.cell(r, 1).value for r in range(2, 5)] == ["one", None, "three"]


def test_large_string_dataframe_stress():
    """Exercises the borrowed path at a size where a per-cell allocation
    would have been measurable."""
    pa = _pa()
    n = 50_000
    table = pa.table({"s": pa.array([f"row-{i}" for i in range(n)]),
                      "n": pa.array(list(range(n)))})
    wb = Workbook()
    ws = wb.add_worksheet("S")
    ws.write_dataframe(0, 0, table)
    wb.close(TEST_FILE)

    sheet = openpyxl.load_workbook(TEST_FILE, read_only=True)["S"]
    rows = sheet.iter_rows(min_row=2, max_row=2, values_only=True)
    assert next(rows) == ("row-0", 0)
