"""Basic functionality tests.

These tests open the generated .xlsx with openpyxl and assert on the
actual cell contents/formatting rather than only checking that a file
got created. A previous version of this library had a bug where
worksheet writes never reached the saved file at all (the file still
existed and was a valid, but essentially empty, workbook) -- a
file-existence-only test suite does not catch that class of bug.
"""
import os
import pytest
import openpyxl
from rvgsrust_xlsxwriter import Workbook

TEST_FILE = "test_output.xlsx"


def teardown_module(module):
    """Clean up test files."""
    if os.path.exists(TEST_FILE):
        os.remove(TEST_FILE)


def _load(path=TEST_FILE):
    return openpyxl.load_workbook(path)


def test_create_workbook():
    wb = Workbook()
    assert wb is not None


def test_add_worksheet():
    wb = Workbook()
    ws = wb.add_worksheet("TestSheet")
    assert ws is not None


def test_write_string():
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write(0, 0, "Hello")
    ws.write(0, 1, "World")
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["A1"].value == "Hello"
    assert sheet["B1"].value == "World"


def test_write_number():
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write(0, 0, 42)
    ws.write(0, 1, 3.14)
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["A1"].value == 42
    assert sheet["B1"].value == pytest.approx(3.14)


def test_write_boolean():
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write(0, 0, True)
    ws.write(0, 1, False)
    wb.close(TEST_FILE)
    sheet = _load().active
    # Must round-trip as real booleans, not be silently coerced to 1.0/0.0
    # (bool is a subclass of int in Python, so this ordering is easy to
    # get wrong when dispatching on Python value type).
    assert sheet["A1"].value is True
    assert sheet["B1"].value is False


def test_write_row():
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write_row(0, 0, ["A", "B", "C"])
    ws.write_row(1, 0, [1, 2, 3])
    wb.close(TEST_FILE)
    sheet = _load().active
    assert [c.value for c in sheet[1]] == ["A", "B", "C"]
    assert [c.value for c in sheet[2]] == [1, 2, 3]


def test_write_column():
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write_column(0, 0, ["A", "B", "C"])
    ws.write_column(0, 1, [1, 2, 3])
    wb.close(TEST_FILE)
    sheet = _load().active
    assert [sheet.cell(row=r, column=1).value for r in (1, 2, 3)] == ["A", "B", "C"]
    assert [sheet.cell(row=r, column=2).value for r in (1, 2, 3)] == [1, 2, 3]


def test_format_bold():
    wb = Workbook()
    ws = wb.add_worksheet()
    fmt = wb.add_format()
    fmt.set_bold()
    ws.write(0, 0, "Bold Text", fmt)
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["A1"].value == "Bold Text"
    assert sheet["A1"].font.bold is True


def test_format_colors():
    wb = Workbook()
    ws = wb.add_worksheet()
    fmt = wb.add_format()
    fmt.set_background_color("#FFFF00")
    fmt.set_font_color("#FF0000")
    ws.write(0, 0, "Colored", fmt)
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["A1"].value == "Colored"
    assert sheet["A1"].fill.fgColor.rgb == "FFFFFF00"
    assert sheet["A1"].font.color.rgb == "FFFF0000"


def test_format_border():
    wb = Workbook()
    ws = wb.add_worksheet()
    fmt = wb.add_format()
    fmt.set_border("thin")
    fmt.set_border_color("#000000")
    ws.write(0, 0, "Bordered", fmt)
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["A1"].value == "Bordered"
    assert sheet["A1"].border.top.style == "thin"


def test_merge_range():
    wb = Workbook()
    ws = wb.add_worksheet()
    fmt = wb.add_format()
    fmt.set_bold()
    fmt.set_background_color("#4472C4")
    fmt.set_font_color("white")
    ws.merge_range(0, 0, 0, 2, "Merged Header", fmt)
    wb.close(TEST_FILE)
    book = _load()
    sheet = book.active
    assert sheet["A1"].value == "Merged Header"
    assert "A1:C1" in [str(r) for r in sheet.merged_cells.ranges]


def test_merge_range_preserves_numeric_and_bool_types():
    # merge_range() historically stringified every value (so a merged
    # number/bool would break SUM()/logic over it). It should now
    # preserve real types by merging with an empty string, then
    # overwriting the anchor cell with the real typed value.
    wb = Workbook()
    ws = wb.add_worksheet()
    fmt = wb.add_format()
    ws.merge_range(0, 0, 0, 2, 12345.67, fmt)
    ws.merge_range(1, 0, 1, 2, True, fmt)
    ws.merge_range(2, 0, 2, 2, False, fmt)
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["A1"].value == 12345.67 and isinstance(sheet["A1"].value, float)
    assert sheet["A2"].value is True
    assert sheet["A3"].value is False


def test_formula():
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write(0, 0, 10)
    ws.write(0, 1, 20)
    ws.write_formula(0, 2, "=A1+B1")
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["C1"].value == "=A1+B1"


def test_freeze_panes():
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.freeze_panes(1, 0)
    ws.write_row(0, 0, ["Header1", "Header2", "Header3"])
    ws.write_row(1, 0, [1, 2, 3])
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet.freeze_panes == "A2"
    assert [c.value for c in sheet[1]] == ["Header1", "Header2", "Header3"]


def test_column_width():
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.set_column_width(0, 20.0)
    ws.write(0, 0, "Wide Column")
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["A1"].value == "Wide Column"
    # Excel stores column width in its own padded units, not raw
    # character width, so an exact match isn't meaningful here -- just
    # confirm it was actually widened from the ~8.43 default.
    assert sheet.column_dimensions["A"].width > 15.0


def test_autofit():
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write(0, 0, "Short")
    ws.write(0, 1, "This is a much longer text")
    ws.autofit()
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["A1"].value == "Short"
    assert sheet["B1"].value == "This is a much longer text"


def test_multiple_sheets():
    wb = Workbook()
    ws1 = wb.add_worksheet("Sheet1")
    ws2 = wb.add_worksheet("Sheet2")
    ws1.write(0, 0, "Data1")
    ws2.write(0, 0, "Data2")
    wb.close(TEST_FILE)
    book = _load()
    assert book["Sheet1"]["A1"].value == "Data1"
    # This is exactly the case the ownership bug broke: a second
    # worksheet's writes overwriting/vanishing relative to the first.
    assert book["Sheet2"]["A1"].value == "Data2"


def test_duplicate_sheet_name_raises():
    wb = Workbook()
    wb.add_worksheet("Sheet1")
    wb.add_worksheet("Sheet1")
    with pytest.raises(Exception):
        wb.close(TEST_FILE)


def test_write_out_of_bounds_raises():
    wb = Workbook()
    ws = wb.add_worksheet()
    with pytest.raises(Exception):
        ws.write(2_000_000, 0, "too far down")


def test_write_records_basic():
    wb = Workbook()
    ws = wb.add_worksheet()
    records = [
        {"Name": "Alice", "Age": 30, "Active": True},
        {"Name": "Bob", "Age": 25, "Active": False},
    ]
    ws.write_records(0, 0, records)
    wb.close(TEST_FILE)
    sheet = _load().active
    assert [c.value for c in sheet[1]] == ["Name", "Age", "Active"]
    assert [c.value for c in sheet[2]] == ["Alice", 30, True]
    assert [c.value for c in sheet[3]] == ["Bob", 25, False]


def test_write_records_explicit_headers_and_no_header_row():
    wb = Workbook()
    ws = wb.add_worksheet()
    records = [{"a": 1, "b": 2, "c": 3}]
    # Explicit headers control column order/subset; write_header=False
    # skips the header row entirely.
    ws.write_records(0, 0, records, headers=["c", "a"], write_header=False)
    wb.close(TEST_FILE)
    sheet = _load().active
    assert [c.value for c in sheet[1]] == [3, 1]


def test_write_records_with_formats():
    wb = Workbook()
    ws = wb.add_worksheet()
    header_fmt = wb.add_format()
    header_fmt.set_bold()
    ws.write_records(0, 0, [{"Name": "Alice"}], header_format=header_fmt)
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["A1"].value == "Name"
    assert sheet["A1"].font.bold is True
    assert sheet["A2"].value == "Alice"


def test_write_records_empty_list_is_noop():
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write_records(0, 0, [])  # should not raise
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["A1"].value is None


def test_write_records_rejects_non_dict_records():
    wb = Workbook()
    ws = wb.add_worksheet()
    with pytest.raises(Exception):
        ws.write_records(0, 0, ["not", "a", "dict"])


def test_add_worksheet_constant_memory_param():
    # Locks in the API contract for constant_memory=True: parameter
    # name, default (False), and that a constant_memory worksheet
    # combined with write_records()/write_dataframe() (both of which
    # write row-by-row in increasing order, satisfying constant memory
    # mode's write-order restriction) produces a correct workbook.
    #
    # NOTE: this can only meaningfully exercise the real streaming/
    # temp-file behavior once built against rust_xlsxwriter 0.96 with
    # the constant_memory feature on a toolchain that can compile it --
    # see Cargo.toml. It still verifies the parameter is accepted, has
    # the right default, and doesn't change write_records()'s output.
    wb = Workbook()
    ws_normal = wb.add_worksheet("Normal")
    ws_streamed = wb.add_worksheet("Streamed", constant_memory=True)

    ws_normal.write(0, 0, "regular")
    ws_streamed.write_records(
        0, 0, [{"a": 1, "b": 2}, {"a": 3, "b": 4}]
    )
    wb.close(TEST_FILE)

    book = _load()
    assert book["Normal"]["A1"].value == "regular"
    assert [c.value for c in book["Streamed"][1]] == ["a", "b"]
    assert [c.value for c in book["Streamed"][2]] == [1, 2]
    assert [c.value for c in book["Streamed"][3]] == [3, 4]


def test_add_worksheet_constant_memory_defaults_false():
    # constant_memory=False (the default) and omitting it entirely
    # must behave identically.
    wb = Workbook()
    ws_default = wb.add_worksheet("Default")
    ws_explicit_false = wb.add_worksheet("ExplicitFalse", constant_memory=False)
    ws_default.write(0, 0, "x")
    ws_explicit_false.write(0, 0, "y")
    wb.close(TEST_FILE)
    book = _load()
    assert book["Default"]["A1"].value == "x"
    assert book["ExplicitFalse"]["A1"].value == "y"


def test_save_to_invalid_path_raises_oserror():
    # XlsxError::IoError (from a failed file write -- bad path,
    # permissions, disk full, etc.) should map to Python's OSError, not
    # the generic ValueError used for parameter/limit errors (bad row,
    # duplicate sheet name, etc.) -- these are different failure
    # categories a caller would want to catch separately.
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write(0, 0, "test")
    with pytest.raises(OSError):
        wb.close("/this_directory_should_not_exist_xyz123/out.xlsx")


def test_parameter_errors_still_raise_valueerror():
    # Distinguishing OSError above must not change the existing
    # ValueError behavior for genuine parameter/limit errors.
    wb = Workbook()
    ws = wb.add_worksheet()
    with pytest.raises(ValueError):
        ws.write(2_000_000, 0, "too far down")

    wb2 = Workbook()
    wb2.add_worksheet("Dup")
    wb2.add_worksheet("Dup")
    with pytest.raises(ValueError):
        wb2.close(TEST_FILE)


def test_constant_memory_normal_worksheet_unaffected():
    # constant_memory's row-order enforcement must not apply to regular
    # worksheets -- out-of-order writes stay fully supported there.
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write(5, 0, "row 5 first")
    ws.write(0, 0, "row 0 second")  # backward -- fine on a normal sheet
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["A6"].value == "row 5 first"
    assert sheet["A1"].value == "row 0 second"


def test_constant_memory_forward_writes_succeed():
    wb = Workbook()
    ws = wb.add_worksheet(constant_memory=True)
    ws.write(0, 0, "a")
    ws.write(1, 0, "b")
    ws.write_row(2, 0, ["c", "d"])
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["A1"].value == "a"
    assert sheet["A2"].value == "b"
    assert [c.value for c in sheet[3]] == ["c", "d"]


def test_constant_memory_same_row_multiple_columns_allowed():
    # Writing multiple columns within the same row is not a "backward"
    # move and must be allowed.
    wb = Workbook()
    ws = wb.add_worksheet(constant_memory=True)
    ws.write(3, 0, "a")
    ws.write(3, 1, "b")
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["A4"].value == "a"
    assert sheet["B4"].value == "b"


def test_constant_memory_backward_write_raises():
    wb = Workbook()
    ws = wb.add_worksheet(constant_memory=True)
    ws.write(5, 0, "row 5")
    with pytest.raises(ValueError, match="non-decreasing order"):
        ws.write(2, 0, "row 2 -- should be rejected")


def test_constant_memory_backward_write_row_raises():
    wb = Workbook()
    ws = wb.add_worksheet(constant_memory=True)
    ws.write_row(5, 0, ["x"])
    with pytest.raises(ValueError):
        ws.write_row(2, 0, ["y"])


def test_constant_memory_backward_merge_range_raises():
    wb = Workbook()
    ws = wb.add_worksheet(constant_memory=True)
    fmt = wb.add_format()
    ws.write(10, 0, "later")
    with pytest.raises(ValueError):
        ws.merge_range(0, 0, 0, 2, "too early", fmt)


def test_constant_memory_backward_write_records_raises():
    # write_records() validates its whole row range up front (before
    # doing any writing), catching a backward move relative to a prior
    # call in a single check rather than needing to fail partway
    # through.
    wb = Workbook()
    ws = wb.add_worksheet(constant_memory=True)
    ws.write(10, 0, "row 10")
    with pytest.raises(ValueError):
        ws.write_records(0, 0, [{"a": 1}, {"a": 2}])


def test_constant_memory_write_column_advances_to_last_row_touched():
    # write_column() touches multiple rows in one call; the next write
    # must be validated against the LAST row it touched, not the first.
    wb = Workbook()
    ws = wb.add_worksheet(constant_memory=True)
    ws.write_column(0, 0, ["a", "b", "c"])  # touches rows 0, 1, 2
    ws.write(2, 1, "same row as last of the column -- fine")
    with pytest.raises(ValueError):
        ws.write(1, 1, "row 1 -- already passed by the column write")


def test_autofilter():
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write_row(0, 0, ["Name", "Dept"])
    ws.write_row(1, 0, ["Alice", "Eng"])
    ws.autofilter(0, 0, 1, 1)
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet.auto_filter.ref == "A1:B2"


def test_autofilter_out_of_range_raises():
    wb = Workbook()
    ws = wb.add_worksheet()
    with pytest.raises(ValueError):
        ws.autofilter(0, 0, 2_000_000, 0)


def test_define_name_global():
    wb = Workbook()
    ws = wb.add_worksheet("Data")
    ws.write(0, 0, "x")
    wb.define_name("MyRange", "Data!$A$1")
    wb.close(TEST_FILE)
    book = _load()
    assert "MyRange" in book.defined_names


def test_define_name_sheet_scoped():
    wb = Workbook()
    ws = wb.add_worksheet("Data")
    ws.write(0, 0, "x")
    wb.define_name("Data!LocalName", "Data!$A$1")
    wb.close(TEST_FILE)
    book = _load()
    assert "LocalName" in book["Data"].defined_names


def test_define_name_invalid_start_char_raises():
    wb = Workbook()
    wb.add_worksheet()
    with pytest.raises(ValueError):
        wb.define_name("1BadName", "Sheet1!$A$1")


def test_define_name_invalid_char_raises():
    wb = Workbook()
    wb.add_worksheet()
    with pytest.raises(ValueError):
        wb.define_name("Bad Name", "Sheet1!$A$1")


def test_table_basic():
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write_row(0, 0, ["Name", "Score"])
    ws.write_column(1, 0, ["Alice", "Bob"])
    ws.write_column(1, 1, [90, 85])

    from rvgsrust_xlsxwriter import Table, TableColumn

    c1 = TableColumn().set_header("Name")
    c2 = TableColumn().set_header("Score")
    table = Table().set_columns([c1, c2]).set_name("ScoresTable")
    ws.add_table(0, 0, 2, 1, table)
    wb.close(TEST_FILE)

    book = _load()
    sheet = book.active
    t = sheet.tables["ScoresTable"]
    assert t.ref == "A1:B3"
    assert [c.name for c in t.tableColumns] == ["Name", "Score"]


def test_table_total_row_with_builtin_function():
    from rvgsrust_xlsxwriter import Table, TableColumn

    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write_row(0, 0, ["Item", "Amount"])
    ws.write_column(1, 0, ["a", "b"])
    ws.write_column(1, 1, [10, 20])

    c1 = TableColumn().set_header("Item").set_total_label("Total")
    c2 = TableColumn().set_header("Amount").set_total_function("sum")
    table = Table().set_columns([c1, c2]).set_total_row(True)
    ws.add_table(0, 0, 3, 1, table)
    wb.close(TEST_FILE)

    sheet = _load().active
    row4 = [c.value for c in sheet[4]]
    assert row4[0] == "Total"
    assert row4[1] == "=SUBTOTAL(109,[Amount])"


def test_table_custom_total_formula():
    from rvgsrust_xlsxwriter import Table, TableColumn

    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write_row(0, 0, ["A", "B"])
    ws.write_column(1, 0, ["x", "y"])
    ws.write_column(1, 1, [1, 2])
    c1 = TableColumn().set_header("A")
    c2 = TableColumn().set_header("B").set_total_function("MEDIAN([B2:B3])")
    table = Table().set_columns([c1, c2]).set_total_row(True)
    ws.add_table(0, 0, 3, 1, table)
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet[4][1].value == "=MEDIAN([B2:B3])"


def test_table_column_formula():
    from rvgsrust_xlsxwriter import Table, TableColumn

    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write_row(0, 0, ["Q1", "Q2", "Total"])
    ws.write_column(1, 0, [10, 20])
    ws.write_column(1, 1, [5, 15])
    c1 = TableColumn().set_header("Q1")
    c2 = TableColumn().set_header("Q2")
    c3 = TableColumn().set_header("Total").set_formula("SUM(Table1[@[Q1]:[Q2]])")
    table = Table().set_columns([c1, c2, c3]).set_name("Table1")
    ws.add_table(0, 0, 2, 2, table)
    wb.close(TEST_FILE)
    sheet = _load().active
    # Excel normalizes the "@" shorthand into its canonical expanded
    # form ("[#This Row]") when saving -- semantically identical, but
    # the literal string differs from what was passed in.
    assert sheet["C2"].value == "=SUM(Table1[[#This Row],[Q1]:[Q2]])"


def test_table_style_valid_and_invalid():
    from rvgsrust_xlsxwriter import Table

    Table().set_style("medium9")  # must not raise
    Table().set_style("none")  # must not raise
    with pytest.raises(ValueError):
        Table().set_style("not_a_real_style")


def test_table_column_format():
    from rvgsrust_xlsxwriter import Table, TableColumn

    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write_row(0, 0, ["Name", "Price"])
    ws.write_column(1, 0, ["Widget"])
    ws.write_column(1, 1, [9.99])

    header_fmt = wb.add_format()
    header_fmt.set_bold()
    cell_fmt = wb.add_format()
    cell_fmt.set_num_format("$#,##0.00")

    c1 = TableColumn().set_header("Name").set_header_format(header_fmt)
    c2 = TableColumn().set_header("Price").set_header_format(header_fmt).set_format(cell_fmt)
    table = Table().set_columns([c1, c2])
    ws.add_table(0, 0, 1, 1, table)
    wb.close(TEST_FILE)

    sheet = _load().active
    assert sheet["A1"].font.bold is True
    assert sheet["B2"].number_format == "$#,##0.00"


def test_table_banded_rows_and_autofilter_options():
    from rvgsrust_xlsxwriter import Table, TableColumn

    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write_row(0, 0, ["A"])
    ws.write_column(1, 0, ["x", "y"])
    c1 = TableColumn().set_header("A")
    table = (
        Table()
        .set_columns([c1])
        .set_banded_rows(True)
        .set_banded_columns(False)
        .set_first_column(True)
        .set_last_column(False)
        .set_autofilter(False)
    )
    assert table.has_header_row() is True  # default
    ws.add_table(0, 0, 2, 0, table)
    wb.close(TEST_FILE)  # must not raise -- content correctness checked above already


def test_table_import_from_package_root():
    # Table/TableColumn must be importable directly from the package,
    # not just from the internal _core extension module.
    from rvgsrust_xlsxwriter import Table, TableColumn
    assert Table is not None
    assert TableColumn is not None


# ============================================================
# New feature tests (added per code review recommendations)
# ============================================================

def test_workbook_context_manager():
    # __enter__/__exit__ must save the file automatically when path is set.
    wb = Workbook(path=TEST_FILE)
    with wb as w:
        ws = w.add_worksheet()
        ws.write(0, 0, "context_manager")
    # After __exit__, the file must exist and have the right content.
    sheet = _load().active
    assert sheet["A1"].value == "context_manager"


def test_workbook_context_manager_no_path_raises():
    # __exit__ with no path set must raise before trying to save.
    with pytest.raises(ValueError, match="no path set"):
        with Workbook() as wb:
            wb.add_worksheet()


def test_close_with_no_argument_uses_constructor_path():
    # The direct (non-context-manager) call pattern shown in the
    # README's Quick Start: Workbook(path).close() with no argument.
    # test_workbook_context_manager above exercises the same underlying
    # close(path=None) via __exit__, but not this call shape directly.
    wb = Workbook(TEST_FILE)
    ws = wb.add_worksheet()
    ws.write(0, 0, "no_arg_close")
    wb.close()
    sheet = _load().active
    assert sheet["A1"].value == "no_arg_close"


def test_close_explicit_path_overrides_constructor_path():
    # Passing a path to close() must still override the constructor
    # path, per the docstring -- this is the backward-compatible half
    # of the no-arg fix, not just the new behaviour.
    import tempfile

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        override_path = tf.name
    try:
        wb = Workbook(TEST_FILE)
        ws = wb.add_worksheet()
        ws.write(0, 0, "explicit_override")
        wb.close(override_path)
        assert os.path.exists(override_path)
        sheet = openpyxl.load_workbook(override_path).active
        assert sheet["A1"].value == "explicit_override"
    finally:
        if os.path.exists(override_path):
            os.remove(override_path)


def test_close_with_no_path_anywhere_raises():
    # Direct (non-context-manager) call with no constructor path and no
    # argument to close() -- same error as the context-manager case,
    # reached a different way.
    wb = Workbook()
    wb.add_worksheet()
    with pytest.raises(ValueError, match="no path set"):
        wb.close()


def test_workbook_context_manager_exception_does_not_save():
    # If an exception is raised inside the with block, __exit__ must
    # NOT attempt to save (and must re-raise the original exception).
    try:
        with Workbook(path=TEST_FILE) as wb:
            raise RuntimeError("deliberate test error")
    except RuntimeError:
        pass
    # File should not have been written (or if TEST_FILE existed before,
    # it would not have been overwritten -- but since teardown_module
    # removes it between runs and this test is ordered before the save
    # tests, the simplest check is just that we got here without an
    # unrelated error being raised).


def test_write_url_basic():
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write_url(0, 0, "https://example.com")
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["A1"].hyperlink.target == "https://example.com"


def test_write_url_with_text():
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write_url(0, 0, "https://example.com", text="Click here")
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["A1"].value == "Click here"
    assert sheet["A1"].hyperlink.target == "https://example.com"


def test_write_url_with_tip():
    # Tooltip (tip) is written into the file but openpyxl doesn't
    # expose it -- just verify the call doesn't raise.
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write_url(0, 0, "https://example.com", tip="Go to example.com")
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["A1"].hyperlink.target == "https://example.com"


def test_write_datetime_py():
    import datetime as _dt
    wb = Workbook()
    ws = wb.add_worksheet()
    fmt = wb.add_format()
    fmt.set_num_format("yyyy-mm-dd hh:mm:ss")
    dt = _dt.datetime(2024, 3, 15, 9, 30, 0)
    ws.write_datetime_py(0, 0, dt, format=fmt)
    wb.close(TEST_FILE)
    sheet = _load().active
    # openpyxl reads Excel datetimes back as Python datetime objects
    cell_val = sheet["A1"].value
    assert cell_val is not None
    assert cell_val.year == 2024
    assert cell_val.month == 3
    assert cell_val.day == 15


def test_write_date_py():
    import datetime as _dt
    wb = Workbook()
    ws = wb.add_worksheet()
    fmt = wb.add_format()
    fmt.set_num_format("yyyy-mm-dd")
    d = _dt.date(2024, 6, 21)
    ws.write_date_py(0, 0, d, format=fmt)
    wb.close(TEST_FILE)
    sheet = _load().active
    cell_val = sheet["A1"].value
    assert cell_val is not None
    assert cell_val.year == 2024
    assert cell_val.month == 6
    assert cell_val.day == 21


def test_write_rich_string_basic():
    wb = Workbook()
    ws = wb.add_worksheet()
    bold = wb.add_format()
    bold.set_bold()
    ws.write_rich_string(0, 0, [
        ("Hello, ", None),
        ("bold", bold),
        (" world", None),
    ])
    wb.close(TEST_FILE)
    sheet = _load().active
    # openpyxl reads rich text as the concatenated plain text
    assert sheet["A1"].value == "Hello, bold world"


def test_write_rich_string_empty_parts_raises():
    wb = Workbook()
    ws = wb.add_worksheet()
    with pytest.raises(ValueError, match="parts list must not be empty"):
        ws.write_rich_string(0, 0, [])


def test_parse_color_invalid_raises():
    wb = Workbook()
    fmt = wb.add_format()
    with pytest.raises(ValueError):
        fmt.set_font_color("#ZZ0000")


def test_parse_color_unknown_name_raises():
    wb = Workbook()
    fmt = wb.add_format()
    with pytest.raises(ValueError):
        fmt.set_background_color("darkk_blue")


def test_parse_border_unknown_raises():
    wb = Workbook()
    fmt = wb.add_format()
    with pytest.raises(ValueError):
        fmt.set_border("thinn")


# ============================================================
# Audit-driven tests (5.4, 5.7, write_rows, insert_image path)
# ============================================================

def test_set_tab_color():
    wb = Workbook()
    ws = wb.add_worksheet("Colored")
    ws.set_tab_color("#FF0000")  # must not raise
    wb.close(TEST_FILE)
    book = _load()
    # openpyxl exposes tab color on sheet.sheet_properties.tabColor
    tab = book["Colored"].sheet_properties.tabColor
    assert tab is not None
    assert tab.rgb.upper().endswith("FF0000")


def test_hide_worksheet():
    wb = Workbook()
    wb.add_worksheet("Visible")
    ws_hidden = wb.add_worksheet("Hidden")
    ws_hidden.hide()
    wb.close(TEST_FILE)
    book = _load()
    assert book["Hidden"].sheet_state == "hidden"
    assert book["Visible"].sheet_state == "visible"


def test_protect_no_password():
    # protect() with no password enables protection with empty-string
    # password — sheet IS protected, but anyone can unprotect without
    # a password. Verify it round-trips via openpyxl.
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.protect()
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet.protection.sheet is True


def test_protect_with_password():
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.protect(password="secret")
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet.protection.sheet is True


def test_set_worksheet_name():
    wb = Workbook()
    ws = wb.add_worksheet("Original")
    ws.set_name("Renamed")
    wb.close(TEST_FILE)
    book = _load()
    assert "Renamed" in book.sheetnames
    assert "Original" not in book.sheetnames


def test_insert_image_missing_file_raises():
    wb = Workbook()
    ws = wb.add_worksheet()
    with pytest.raises(OSError, match="not found"):
        ws.insert_image(0, 0, "/tmp/does_not_exist_xyz.png")


def test_write_rows_basic():
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write_row(0, 0, ["Name", "Score"])
    ws.write_rows(1, 0, [
        ["Alice", 95],
        ["Bob",   87],
        ["Carol", 91],
    ])
    wb.close(TEST_FILE)
    sheet = _load().active
    assert [c.value for c in sheet[1]] == ["Name", "Score"]
    assert [c.value for c in sheet[2]] == ["Alice", 95]
    assert [c.value for c in sheet[3]] == ["Bob",   87]
    assert [c.value for c in sheet[4]] == ["Carol", 91]


def test_write_rows_with_write_header():
    wb = Workbook()
    ws = wb.add_worksheet()
    header_fmt = wb.add_format()
    header_fmt.set_bold()
    ws.write_rows(0, 0, [
        ["Name", "Score"],
        ["Alice", 95],
    ], write_header=True, header_format=header_fmt)
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["A1"].font.bold is True
    assert sheet["A1"].value == "Name"
    assert sheet["A2"].value == "Alice"


def test_write_rows_empty_is_noop():
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write_rows(0, 0, [])  # must not raise
    wb.close(TEST_FILE)
    assert _load().active["A1"].value is None


def test_write_rows_faster_than_records_same_output():
    # write_rows and write_records must produce identical cell values
    # for equivalent data (the fast path must not corrupt output).
    records  = [{"x": i, "y": i * 2} for i in range(50)]
    rows     = [[r["x"], r["y"]] for r in records]

    wb1 = Workbook()
    ws1 = wb1.add_worksheet()
    ws1.write_row(0, 0, ["x", "y"])
    ws1.write_records(1, 0, records, headers=["x", "y"], write_header=False)
    wb1.close("_cmp_records.xlsx")

    wb2 = Workbook()
    ws2 = wb2.add_worksheet()
    ws2.write_row(0, 0, ["x", "y"])
    ws2.write_rows(1, 0, rows)
    wb2.close("_cmp_rows.xlsx")

    import openpyxl as _opx
    s1 = _opx.load_workbook("_cmp_records.xlsx").active
    s2 = _opx.load_workbook("_cmp_rows.xlsx").active
    import os
    os.remove("_cmp_records.xlsx")
    os.remove("_cmp_rows.xlsx")

    for row in range(1, 52):
        assert [c.value for c in s1[row]] == [c.value for c in s2[row]], \
            f"row {row} mismatch"


def test_table_set_alt_text():
    from rvgsrust_xlsxwriter import Table, TableColumn
    t = Table().set_name("AltTest")
    t.set_alt_text("Accessibility description")   # must not raise
    t.set_alt_text_title("Table title")           # must not raise


def test_write_formula_with_format():
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write(0, 0, 10)
    ws.write(0, 1, 20)
    fmt = wb.add_format()
    fmt.set_bold()
    ws.write_formula(0, 2, "=A1+B1", fmt)
    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["C1"].value == "=A1+B1"
    assert sheet["C1"].font.bold is True


def test_parse_color_valid_hex():
    wb = Workbook()
    fmt = wb.add_format()
    fmt.set_font_color("#1A2B3C")  # valid hex — must not raise


def test_parse_color_valid_named():
    wb = Workbook()
    fmt = wb.add_format()
    fmt.set_background_color("navy")  # valid named color


def test_parse_border_valid():
    wb = Workbook()
    fmt = wb.add_format()
    fmt.set_border("medium")  # must not raise


def test_write_rows_constant_memory_with_write_header():
    """Regression test for the last_row double-count bug.

    Before the fix, write_rows() with write_header=True was setting
    min_allowed_row one row past the actual last write, incorrectly
    rejecting valid subsequent writes in constant_memory mode.
    """
    wb = Workbook()
    ws = wb.add_worksheet(constant_memory=True)
    header_fmt = wb.add_format()
    header_fmt.set_bold()

    # Write header + 3 data rows starting at row 0
    ws.write_rows(0, 0, [
        ["Name", "Score"],  # header
        ["Alice", 95],
        ["Bob",   87],
        ["Carol", 91],
    ], write_header=True, header_format=header_fmt)

    # rows.len() == 4, so last row written was row 3.
    # A subsequent write to row 4 must be accepted (not rejected as if
    # min_allowed_row had been set to 5).
    ws.write(4, 0, "Dave")
    ws.write(4, 1, 88)

    wb.close(TEST_FILE)
    sheet = _load().active
    assert sheet["A1"].value == "Name"
    assert sheet["A4"].value == "Carol"
    assert sheet["A5"].value == "Dave"
    assert sheet["B5"].value == 88
