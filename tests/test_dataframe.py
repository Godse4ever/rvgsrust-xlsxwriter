"""DataFrame support tests."""
import pytest
from rvgsrust_xlsxwriter import Workbook
from rvgsrust_xlsxwriter.dataframe import write_polars_dataframe, write_pandas_dataframe

TEST_FILE = "test_df_output.xlsx"


try:
    import polars as pl
    HAS_POLARS = True
except ImportError:
    HAS_POLARS = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


@pytest.mark.skipif(not HAS_POLARS, reason="Polars not installed")
def test_polars_dataframe():
    df = pl.DataFrame({
        "Name": ["Alice", "Bob", "Charlie"],
        "Age": [30, 25, 35],
        "Salary": [50000.0, 60000.0, 70000.0],
    })

    wb = Workbook()
    ws = wb.add_worksheet()

    header_fmt = wb.add_format()
    header_fmt.set_bold()
    header_fmt.set_background_color("#4472C4")
    header_fmt.set_font_color("white")

    write_polars_dataframe(ws, df, row=0, col=0, header_format=header_fmt)
    ws.autofit()
    wb.close(TEST_FILE)


@pytest.mark.skipif(not HAS_PANDAS, reason="Pandas not installed")
def test_pandas_dataframe():
    df = pd.DataFrame({
        "Product": ["Widget", "Gadget", "Thingama"],
        "Price": [9.99, 19.99, 29.99],
        "Stock": [100, 50, 25],
    })

    wb = Workbook()
    ws = wb.add_worksheet()

    header_fmt = wb.add_format()
    header_fmt.set_bold()
    header_fmt.set_background_color("#70AD47")
    header_fmt.set_font_color("white")

    write_pandas_dataframe(ws, df, row=0, col=0, header_format=header_fmt)
    ws.autofit()
    wb.close(TEST_FILE)


try:
    import pyarrow as pa
    HAS_PYARROW = True
except ImportError:
    HAS_PYARROW = False

import openpyxl


def _load(path=TEST_FILE):
    return openpyxl.load_workbook(path)


@pytest.mark.skipif(not HAS_PYARROW, reason="PyArrow not installed")
def test_write_dataframe_pyarrow_table():
    table = pa.table({
        "id": pa.array([1, 2, 3], type=pa.int64()),
        "name": pa.array(["Alice", "Bob", "Carol"], type=pa.string()),
        "score": pa.array([95.5, 88.2, 71.0], type=pa.float64()),
        "passed": pa.array([True, False, True], type=pa.bool_()),
    })
    wb = Workbook()
    ws = wb.add_worksheet()
    fmt = wb.add_format()
    fmt.set_bold()
    ws.write_dataframe(0, 0, table, header_format=fmt)
    wb.close(TEST_FILE)

    sheet = _load().active
    assert [c.value for c in sheet[1]] == ["id", "name", "score", "passed"]
    assert sheet["A1"].font.bold is True
    assert [c.value for c in sheet[2]] == [1, "Alice", 95.5, True]


@pytest.mark.skipif(not HAS_PYARROW, reason="PyArrow not installed")
def test_write_dataframe_nulls():
    table = pa.table({
        "a": pa.array([1, None, 3], type=pa.int64()),
        "b": pa.array(["x", "y", None], type=pa.string()),
    })
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write_dataframe(0, 0, table)
    wb.close(TEST_FILE)

    sheet = _load().active
    assert sheet["A3"].value is None  # the None in column a
    assert sheet["B4"].value is None  # the None in column b


@pytest.mark.skipif(not HAS_PANDAS, reason="Pandas not installed")
def test_write_dataframe_pandas_large_utf8():
    # Pandas commonly exports string columns as Arrow LargeUtf8, not
    # Utf8 -- this must work too, not just pyarrow's default Utf8.
    df = pd.DataFrame({"x": [1, 2, 3], "y": ["a", "b", "c"]})
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write_dataframe(0, 0, df)
    wb.close(TEST_FILE)

    sheet = _load().active
    assert [c.value for c in sheet[1]] == ["x", "y"]
    assert [c.value for c in sheet[2]] == [1, "a"]


@pytest.mark.skipif(not HAS_PYARROW, reason="PyArrow not installed")
def test_write_dataframe_unsupported_type_raises():
    # int32 used to stand in for "unsupported" here; it is supported as of
    # the extended Arrow types work, so this uses binary, which is not.
    table = pa.table({"d": pa.array([b"x", b"y"], type=pa.binary())})
    wb = Workbook()
    ws = wb.add_worksheet()
    with pytest.raises(TypeError):
        ws.write_dataframe(0, 0, table)


def test_write_dataframe_rejects_non_dataframe():
    wb = Workbook()
    ws = wb.add_worksheet()
    with pytest.raises(TypeError):
        ws.write_dataframe(0, 0, {"not": "a dataframe"})


# --- Regression tests for the unsafe Arrow PyCapsule handling ---
#
# write_dataframe() takes ownership of a C-level ArrowArrayStream
# struct out of a PyCapsule (record_batches_from_arrow() in
# src/lib.rs), which involves an unsafe ptr::read plus manually
# nulling the source struct's release field so the capsule's own
# destructor doesn't double-release the same underlying data (see the
# detailed comment there for the full Arrow C Data Interface
# reasoning). That's the highest-risk code in this crate: a mistake
# there is a potential double-free or use-after-free, not just a wrong
# answer. These tests exercise it repeatedly and across varied shapes
# to have some chance of catching a memory-safety regression, though
# a clean run here is not a substitute for testing with a sanitizer
# (ASAN/valgrind) on a real build -- pytest can't detect memory
# corruption that doesn't happen to crash or corrupt something we
# assert on.


@pytest.mark.skipif(not HAS_PYARROW, reason="PyArrow not installed")
def test_write_dataframe_repeated_calls_no_crash():
    table = pa.table({
        "id": pa.array(list(range(500)), type=pa.int64()),
        "value": pa.array([float(i) * 1.5 for i in range(500)], type=pa.float64()),
    })
    # Repeated capsule take-ownership cycles in one process -- a
    # double-free from the release-field bug would tend to show up as
    # a crash somewhere in this loop, not necessarily the first call.
    for _ in range(20):
        wb = Workbook()
        ws = wb.add_worksheet()
        ws.write_dataframe(0, 0, table)
        wb.close(TEST_FILE)
    # If we got here without segfaulting, that's the main signal this
    # test can give. Also check the last write actually has content.
    sheet = _load().active
    assert sheet["A2"].value == 0
    assert sheet["B2"].value == 0.0


@pytest.mark.skipif(not HAS_PYARROW, reason="PyArrow not installed")
def test_write_dataframe_multiple_worksheets_same_workbook():
    # Exercises taking ownership of two separate capsules (two
    # separate __arrow_c_stream__() calls) within the same workbook,
    # writing to different worksheets.
    table_a = pa.table({"x": pa.array([1, 2], type=pa.int64())})
    table_b = pa.table({"y": pa.array(["p", "q"], type=pa.string())})
    wb = Workbook()
    ws_a = wb.add_worksheet("A")
    ws_b = wb.add_worksheet("B")
    ws_a.write_dataframe(0, 0, table_a)
    ws_b.write_dataframe(0, 0, table_b)
    wb.close(TEST_FILE)
    book = _load()
    assert book["A"]["A2"].value == 1
    assert book["B"]["A2"].value == "p"


@pytest.mark.skipif(not HAS_PYARROW, reason="PyArrow not installed")
def test_write_dataframe_empty_table():
    # Zero rows: the capsule/stream is still created and consumed even
    # though there's no data to iterate -- makes sure the ownership
    # handling doesn't assume at least one batch exists.
    table = pa.table({"x": pa.array([], type=pa.int64())})
    wb = Workbook()
    ws = wb.add_worksheet()
    ws.write_dataframe(0, 0, table)  # should not raise
    wb.close(TEST_FILE)
    sheet = _load().active
    # No rows to write, but this shouldn't crash and shouldn't produce
    # spurious content either.
    assert sheet["A1"].value is None


def test_write_dataframe_polars_utf8view():
    """Polars >= 1.0 uses Utf8View (StringView) as default string dtype.
    This previously raised TypeError; must now succeed."""
    try:
        import polars as pl
    except ImportError:
        pytest.skip("Polars not installed")

    import pyarrow as pa
    # Force Utf8View by using the native polars string type (default since 1.0)
    df = pl.DataFrame({
        "name": pl.Series(["Alice", "Bob", "Carol"], dtype=pl.String),
        "score": pl.Series([1, 2, 3], dtype=pl.Int64),
    })
    # Convert to arrow — this will produce Utf8View for string columns
    arrow_table = df.to_arrow()
    from rvgsrust_xlsxwriter import Workbook
    import tempfile, os
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        wb = Workbook()
        ws = wb.add_worksheet()
        ws.write_dataframe(0, 0, arrow_table)  # must not raise
        wb.close(path)
        import openpyxl
        sheet = openpyxl.load_workbook(path).active
        assert sheet["A2"].value == "Alice"
        assert sheet["A3"].value == "Bob"
    finally:
        if os.path.exists(path):
            os.remove(path)


# ---------------------------------------------------------------------
# Extended Arrow type coverage (integer widths, float32, dates,
# timestamps). See src/lib.rs SUPPORTED_ARROW_TYPES.
# ---------------------------------------------------------------------

# pa / HAS_PYARROW are already defined near the top of this module.
requires_pyarrow = pytest.mark.skipif(not HAS_PYARROW, reason="pyarrow not installed")


def _roundtrip(table):
    """Write an arrow table and hand back the loaded openpyxl sheet."""
    import openpyxl
    import tempfile
    import os

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        wb = Workbook()
        ws = wb.add_worksheet()
        ws.write_dataframe(0, 0, table)
        wb.close(path)
        return openpyxl.load_workbook(path).active
    finally:
        if os.path.exists(path):
            os.remove(path)


@requires_pyarrow
def test_arrow_narrow_integer_widths():
    """int8/16/32 and uint8/16/32/64 all land as numbers, not TypeError."""
    table = pa.table({
        "i8": pa.array([-128, 0, 127], type=pa.int8()),
        "i16": pa.array([-32768, 0, 32767], type=pa.int16()),
        "i32": pa.array([-2147483648, 0, 2147483647], type=pa.int32()),
        "u8": pa.array([0, 128, 255], type=pa.uint8()),
        "u16": pa.array([0, 1000, 65535], type=pa.uint16()),
        "u32": pa.array([0, 70000, 4294967295], type=pa.uint32()),
        "u64": pa.array([0, 1, 9007199254740992], type=pa.uint64()),
    })
    sheet = _roundtrip(table)
    assert sheet["A1"].value == "i8"
    assert sheet["A2"].value == -128
    assert sheet["A4"].value == 127
    assert sheet["C2"].value == -2147483648
    assert sheet["F4"].value == 4294967295
    # 2^53 is the largest integer f64 represents exactly.
    assert sheet["G4"].value == 9007199254740992


@requires_pyarrow
def test_arrow_float32():
    """f32 widens to f64; exactly-representable values survive intact."""
    table = pa.table({"x": pa.array([1.5, -2.25, 0.0], type=pa.float32())})
    sheet = _roundtrip(table)
    assert sheet["A2"].value == 1.5
    assert sheet["A3"].value == -2.25
    assert sheet["A4"].value == 0.0


@requires_pyarrow
def test_arrow_date32_renders_as_date_not_serial():
    """The regression this feature exists for: a date column must come
    back as a date, not as the integer 45123."""
    import datetime

    table = pa.table({
        "d": pa.array(
            [datetime.date(2023, 7, 14), datetime.date(1970, 1, 1)],
            type=pa.date32(),
        )
    })
    sheet = _roundtrip(table)
    got = sheet["A2"].value
    assert not isinstance(got, (int, float)), f"date wrote as raw serial: {got!r}"
    assert got == datetime.datetime(2023, 7, 14)
    # Unix epoch == Excel serial 25569.
    assert sheet["A3"].value == datetime.datetime(1970, 1, 1)
    assert sheet["A2"].number_format == "yyyy-mm-dd"


@requires_pyarrow
def test_arrow_date64():
    import datetime

    table = pa.table({
        "d": pa.array([datetime.date(2024, 2, 29)], type=pa.date64())
    })
    sheet = _roundtrip(table)
    assert sheet["A2"].value == datetime.datetime(2024, 2, 29)


@requires_pyarrow
@pytest.mark.parametrize("unit", ["s", "ms", "us", "ns"])
def test_arrow_timestamp_all_units(unit):
    import datetime

    expected = datetime.datetime(2023, 6, 18, 17, 8, 28)
    table = pa.table({"t": pa.array([expected], type=pa.timestamp(unit))})
    sheet = _roundtrip(table)
    got = sheet["A2"].value
    assert isinstance(got, datetime.datetime), f"timestamp[{unit}] wrote as {got!r}"
    # f64 serials carry roughly microsecond resolution at modern dates,
    # so compare with a tolerance rather than demanding exact equality.
    assert abs((got - expected).total_seconds()) < 0.001


@pytest.mark.skipif(not HAS_PANDAS, reason="Pandas not installed")
def test_pandas_default_datetime64_ns_column():
    """pandas' default datetime dtype is datetime64[ns]; this is the
    single most common real-world case and must not raise."""
    import datetime

    df = pd.DataFrame({
        "when": pd.to_datetime(["2023-01-15", "2024-12-31"]),
        "what": ["a", "b"],
    })
    sheet = _roundtrip(df)
    assert sheet["A2"].value == datetime.datetime(2023, 1, 15)
    assert sheet["A3"].value == datetime.datetime(2024, 12, 31)
    assert sheet["B2"].value == "a"


@requires_pyarrow
def test_arrow_temporal_nulls_are_blank():
    table = pa.table({
        "d": pa.array([None], type=pa.date32()),
        "t": pa.array([None], type=pa.timestamp("us")),
        "i": pa.array([None], type=pa.int32()),
    })
    sheet = _roundtrip(table)
    assert sheet["A2"].value is None
    assert sheet["B2"].value is None
    assert sheet["C2"].value is None


@requires_pyarrow
def test_tz_aware_timestamp_warns_and_writes_utc():
    import datetime

    table = pa.table({
        "t": pa.array(
            [datetime.datetime(2023, 6, 18, 17, 8, 28, tzinfo=datetime.timezone.utc)],
            type=pa.timestamp("us", tz="UTC"),
        )
    })
    with pytest.warns(UserWarning, match="timezone-aware"):
        sheet = _roundtrip(table)
    got = sheet["A2"].value
    assert abs((got - datetime.datetime(2023, 6, 18, 17, 8, 28)).total_seconds()) < 0.001


@requires_pyarrow
def test_pre_1900_date_raises_valueerror_naming_column():
    """Excel cannot represent dates before 1900; the error should say
    which column and row rather than surfacing a bare serial number."""
    import datetime

    table = pa.table({
        "birth": pa.array([datetime.date(1850, 3, 1)], type=pa.date32())
    })
    with pytest.raises(ValueError) as exc:
        _roundtrip(table)
    assert "birth" in str(exc.value)


@requires_pyarrow
def test_still_unsupported_type_raises_typeerror():
    import decimal

    table = pa.table({
        "amount": pa.array([decimal.Decimal("1.50")], type=pa.decimal128(10, 2))
    })
    with pytest.raises(TypeError) as exc:
        _roundtrip(table)
    msg = str(exc.value)
    assert "amount" in msg
    assert "date32" in msg  # the supported-types list is included


# ---------------------------------------------------------------------
# column_formats: the actual bug this parameter exists to fix is that a
# COLUMN-scoped format (set_column_format / set_column_range_format)
# loses to a cell's own number format under OOXML precedence, so a
# border silently vanishes on any date/datetime column. These tests
# assert the merge happens per cell, not that a column_formats argument
# merely exists.
# ---------------------------------------------------------------------

def _border_count_in_styles(path):
    """Count <border> child elements inside <borders> in styles.xml.
    A shared Format instance passed for every column should still
    produce one border definition, not one per column."""
    import zipfile
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(path) as z:
        xml_bytes = z.read("xl/styles.xml")
    root = ET.fromstring(xml_bytes)
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    borders = root.find("m:borders", ns)
    # The default/empty border (index 0, no sides set) is always present,
    # so a single real border definition on top of that is 2 elements,
    # not 1 -- assert against that baseline rather than a bare count.
    return len(borders.findall("m:border", ns))


@requires_pyarrow
def test_write_dataframe_column_formats_merges_not_overwrites():
    """The regression this feature exists for: a border applied via
    column_formats must survive on a date column, which still carries
    its own numFmtId. Before the fix, a column-scoped workaround
    (set_column_format) silently dropped the border on exactly this
    column, with no error."""
    import datetime
    import openpyxl.utils

    table = pa.table({
        "q_int": pa.array([1, 2], type=pa.int64()),
        "q_float": pa.array([1.5, 2.5], type=pa.float64()),
        "q_str": pa.array(["a", "b"], type=pa.string()),
        "q_date": pa.array(
            [datetime.date(2026, 1, 1), datetime.date(2026, 1, 2)], type=pa.date32()
        ),
    })
    wb = Workbook()
    ws = wb.add_worksheet()
    border_fmt = wb.add_format()
    border_fmt.set_border("thin")
    border_fmt.set_border_color("#BFBFBF")
    # The same Format instance for every column, deliberately: the
    # write side must not require one Format per column to get one
    # border definition in styles.xml.
    ws.write_dataframe(
        0, 0, table, write_header=True,
        column_formats={c: border_fmt for c in table.column_names},
    )
    wb.close(TEST_FILE)

    sheet = _load().active
    n_cols = len(table.column_names)
    for col_idx in range(1, n_cols + 1):
        cell = sheet.cell(row=2, column=col_idx)
        assert cell.border.top.style == "thin", (
            f"column {table.column_names[col_idx - 1]!r} lost its border "
            f"(border={cell.border.top.style!r})"
        )

    # q_date is the 4th column -- must still carry a date number format,
    # not just "General", or the border came at the cost of the date
    # rendering as a raw serial again.
    date_cell = sheet.cell(row=2, column=n_cols)
    assert date_cell.number_format not in ("General", None), (
        f"q_date lost its date number format: {date_cell.number_format!r}"
    )
    assert date_cell.value == datetime.datetime(2026, 1, 1)

    # One shared Format instance across 4 columns of 3 different dtypes
    # (int/float, str, date) should not produce 4 separate border
    # definitions in styles.xml.
    n_borders = _border_count_in_styles(TEST_FILE)
    assert n_borders <= 3, (
        f"expected at most a couple of border definitions (default + "
        f"shared), got {n_borders} -- looks like the format wasn't "
        f"deduped across columns"
    )


@requires_pyarrow
def test_write_dataframe_column_formats_constant_memory():
    """Same as the merge test above, but with constant_memory=True --
    column_formats must work in the streaming path too, since formats
    are applied at write time before rows are flushed."""
    import datetime

    table = pa.table({
        "q_str": pa.array(["a", "b"], type=pa.string()),
        "q_date": pa.array(
            [datetime.date(2026, 1, 1), datetime.date(2026, 1, 2)], type=pa.date32()
        ),
    })
    wb = Workbook()
    ws = wb.add_worksheet(constant_memory=True)
    border_fmt = wb.add_format()
    border_fmt.set_border("thin")
    ws.write_dataframe(
        0, 0, table, write_header=True,
        column_formats={c: border_fmt for c in table.column_names},
    )
    wb.close(TEST_FILE)

    sheet = _load().active
    assert sheet.cell(row=2, column=2).border.top.style == "thin"
    assert sheet.cell(row=2, column=2).number_format not in ("General", None)


@requires_pyarrow
def test_write_dataframe_column_formats_unknown_key_raises():
    table = pa.table({"a": pa.array([1, 2], type=pa.int64())})
    wb = Workbook()
    ws = wb.add_worksheet()
    fmt = wb.add_format()
    fmt.set_bold()
    with pytest.raises(ValueError) as exc:
        ws.write_dataframe(0, 0, table, column_formats={"nope": fmt})
    assert "nope" in str(exc.value)


@requires_pyarrow
def test_write_dataframe_column_formats_none_is_unchanged():
    """Omitting column_formats must be byte-identical to today's
    behaviour -- this is a pure addition, not a behaviour change for
    existing callers."""
    table = pa.table({"a": pa.array([1, 2], type=pa.int64())})
    wb1 = Workbook()
    ws1 = wb1.add_worksheet()
    ws1.write_dataframe(0, 0, table)
    wb1.close(TEST_FILE)
    with open(TEST_FILE, "rb") as f:
        bytes_without_kwarg = f.read()

    wb2 = Workbook()
    ws2 = wb2.add_worksheet()
    ws2.write_dataframe(0, 0, table, column_formats=None)
    wb2.close(TEST_FILE)
    with open(TEST_FILE, "rb") as f:
        bytes_with_explicit_none = f.read()

    assert bytes_without_kwarg == bytes_with_explicit_none
