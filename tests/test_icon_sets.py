"""Conditional format icon set tests.

XML attribute values verified against actual rust_xlsxwriter 0.98.2
source (conditional_format.rs's Display impl for
ConditionalFormatIconType, and the rule()/write_icon_set() XML writer)
before writing these assertions -- notably, ThreeSymbols maps to the
OOXML preset "3Symbols2" and ThreeSymbolsCircled maps to "3Symbols",
reversed from what the names would suggest if guessed. Also: the
iconSet attribute is omitted entirely when the type is the default
(ThreeTrafficLights), so tests use a non-default type to actually
exercise the setter.
"""
import os
import tempfile
import zipfile

import openpyxl
import pytest

from rvgsrust_xlsxwriter import (
    ConditionalFormatCustomIcon,
    ConditionalFormatIconSet,
    Workbook,
)


def _apply(build):
    """Write a small numeric grid, run build(ws), return sheet1.xml."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        wb = Workbook()
        ws = wb.add_worksheet()
        for r in range(5):
            ws.write(r, 0, r)
        build(ws)
        wb.close(path)
        with zipfile.ZipFile(path) as z:
            return z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    finally:
        if os.path.exists(path):
            os.remove(path)


# ---------------------------- basic ----------------------------


def test_default_icon_type_omits_iconset_attribute():
    # ThreeTrafficLights is upstream's own default -- confirmed the
    # writer omits the iconSet="..." attribute entirely in that case,
    # to keep the XML compact.
    def build(ws):
        cf = ConditionalFormatIconSet()
        ws.add_conditional_format(0, 0, 4, 0, cf)

    sheet = _apply(build)
    assert "<iconSet" in sheet
    assert "iconSet=" not in sheet


def test_non_default_icon_type_sets_attribute():
    def build(ws):
        cf = ConditionalFormatIconSet()
        cf.set_icon_type("four_arrows")
        ws.add_conditional_format(0, 0, 4, 0, cf)

    sheet = _apply(build)
    assert 'iconSet="4Arrows"' in sheet


def test_icon_type_name_mapping_not_guessed():
    # The two icon types whose OOXML preset name is reversed from what
    # the Python name would suggest -- pinned explicitly since it would
    # be an easy mistake to get backwards.
    def build_symbols(ws):
        cf = ConditionalFormatIconSet()
        cf.set_icon_type("three_symbols")
        ws.add_conditional_format(0, 0, 4, 0, cf)

    def build_circled(ws):
        cf = ConditionalFormatIconSet()
        cf.set_icon_type("three_symbols_circled")
        ws.add_conditional_format(0, 0, 4, 0, cf)

    assert 'iconSet="3Symbols2"' in _apply(build_symbols)
    assert 'iconSet="3Symbols"' in _apply(build_circled)


def test_invalid_icon_type_raises():
    cf = ConditionalFormatIconSet()
    with pytest.raises(ValueError):
        cf.set_icon_type("three_smileys")


# ---------------------------- reverse / icons-only ----------------------------


def test_reverse_icons_sets_attribute():
    def build(ws):
        cf = ConditionalFormatIconSet()
        cf.reverse_icons(True)
        ws.add_conditional_format(0, 0, 4, 0, cf)

    sheet = _apply(build)
    assert 'reverse="1"' in sheet


def test_show_icons_only_sets_show_value_zero():
    def build(ws):
        cf = ConditionalFormatIconSet()
        cf.show_icons_only(True)
        ws.add_conditional_format(0, 0, 4, 0, cf)

    sheet = _apply(build)
    assert 'showValue="0"' in sheet


def test_defaults_omit_reverse_and_show_value():
    def build(ws):
        cf = ConditionalFormatIconSet()
        ws.add_conditional_format(0, 0, 4, 0, cf)

    sheet = _apply(build)
    assert "reverse=" not in sheet
    assert "showValue=" not in sheet


# ---------------------------- custom icons ----------------------------


def test_custom_icons_produce_valid_file():
    def build(ws):
        icons = [
            ConditionalFormatCustomIcon(),
            ConditionalFormatCustomIcon(),
            ConditionalFormatCustomIcon(),
        ]
        icons[0].set_rule("percent", 0)
        icons[1].set_rule("percent", 33)
        icons[1].set_icon_type("four_histograms", 0)
        icons[2].set_rule("percent", 67)
        icons[2].set_icon_type("five_boxes", 4)

        cf = ConditionalFormatIconSet()
        cf.set_icon_type("three_traffic_lights")
        cf.set_icons(icons)
        ws.add_conditional_format(0, 0, 4, 0, cf)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        path = tf.name
    try:
        wb = Workbook()
        ws = wb.add_worksheet()
        for r in range(5):
            ws.write(r, 0, r)
        build(ws)
        wb.close(path)
        sheet = openpyxl.load_workbook(path).active
        assert sheet["A1"].value == 0
    finally:
        if os.path.exists(path):
            os.remove(path)


def test_custom_icon_index_out_of_range_raises():
    icon = ConditionalFormatCustomIcon()
    with pytest.raises(ValueError):
        icon.set_icon_type("three_arrows", 3)  # valid range is 0-2


def test_custom_icon_index_in_range_does_not_raise():
    icon = ConditionalFormatCustomIcon()
    icon.set_icon_type("three_arrows", 2)  # must not raise
    icon.set_icon_type("five_boxes", 4)  # must not raise


def test_custom_icon_no_icon_and_greater_than():
    icon = ConditionalFormatCustomIcon()
    icon.set_no_icon(True)
    icon.set_greater_than(True)  # must not raise


# ---------------------------- common CF settings ----------------------------


def test_multi_range_and_stop_if_true():
    def build(ws):
        cf = ConditionalFormatIconSet()
        cf.set_multi_range("A1:A5,C1:C5")
        cf.set_stop_if_true(True)
        ws.add_conditional_format(0, 0, 4, 0, cf)

    sheet = _apply(build)
    assert 'stopIfTrue="1"' in sheet


def test_add_conditional_format_rejects_non_cf_object():
    wb = Workbook()
    ws = wb.add_worksheet()
    with pytest.raises(TypeError):
        ws.add_conditional_format(0, 0, 4, 0, "not a conditional format")
